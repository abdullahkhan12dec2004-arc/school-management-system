"""
app.py - Flask Application Routes
School Management System with Role-Based Access
"""

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify

from flask import request, redirect, url_for, flash, session, send_file
from werkzeug.utils import secure_filename
from functools import wraps
import datetime
import os
import psycopg2
from database import get_db, fetchall_dict, fetchone_dict, hash_password, init_db
from reports import reports_bp
import pandas as pd
import hashlib

app = Flask(__name__)
app.secret_key = 'school_mgmt_secret_2024'
app.register_blueprint(reports_bp)
UPLOAD_FOLDER = 'static/uploads/logos'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
# app.py mein import
from models import (
    School, Student, StudentParent, StudentSibling,
    Teacher, TeacherDocument, FeeCollection,
    StudentAttendance, TeacherAttendance, Class
)




def get_active_school_id():
    if session.get('role') == 'school_admin':
        return session.get('school_id')
    return session.get('active_school_id') or session.get('school_id')


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ========== DECORATORS ==========
def school_admin_only_required(f):
    """Sirf School Admin access kar sakta hai — Super Admin (admin) nahi."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'school_admin':
            flash('Yeh option sirf School Admin ke liye hai', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Pehle login karein', 'errordef school_admin_only_required(f):
    """Sirf School Admin access kar sakta hai — Super Admin (admin) nahi."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'school_admin':
            flash('Yeh option sirf School Admin ke liye hai', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated')
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Sirf Admin access kar sakta hai', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)

    return decorated


def school_admin_or_super_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        role = session.get('role')
        if role not in ['admin', 'school_admin']:
            flash('Sirf Admin ya School Admin access kar sakta hai', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def teacher_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') not in ['admin', 'school_admin', 'teacher']:
            flash('Aapko is page ka access nahi hai', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)

    return decorated


ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


def get_col(row_dict, *keys, default=''):
    """
    Multiple column name variants try karta hai.
    Case-insensitive + strip + NaN safe.
    """
    import pandas as pd
    lower_map = {str(k).lower().strip(): v for k, v in row_dict.items()}

    for key in keys:
        val = row_dict.get(key)
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            s = str(val).strip()
            if s and s.lower() not in ('nan', 'none', ''):
                return s
        val = lower_map.get(str(key).lower().strip())
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            s = str(val).strip()
            if s and s.lower() not in ('nan', 'none', ''):
                return s
    return default


def is_row_empty(row_dict):
    """
    True agar row mein koi bhi meaningful value nahi hai.
    """
    import pandas as pd
    skip_keys = {'school id', 'school_id', 'schoolid'}
    for k, val in row_dict.items():
        if str(k).lower().strip() in skip_keys:
            continue
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            s = str(val).strip()
            if s and s.lower() not in ('nan', 'none', ''):
                return False
    return True

def normalize_df_columns(df):
    df.columns = [str(col).strip().rstrip('*').strip() for col in df.columns]
    return df

def validate_classes_row(row_dict, school_id):
    errors = []
    class_name = get_col(row_dict,
                         'Class Name', 'class_name', 'ClassName',
                         'CLASS NAME', 'classname', 'class name')
    section    = get_col(row_dict,
                         'Section', 'section', 'SECTION')

    if not class_name:
        errors.append("Class Name is required")
    if not section:
        errors.append("Section is required")
    return errors


def validate_teacher_row(row_dict, conn):
    errors = []

    full_name    = get_col(row_dict, 'Full Name',    'full_name',    'FullName',    'FULL NAME')
    username     = get_col(row_dict, 'Username',     'username',     'USERNAME')
    password     = get_col(row_dict, 'Password',     'password',     'PASSWORD')
    phone        = get_col(row_dict, 'Phone',        'phone',        'PHONE',       'Phone No', 'phone_no')
    joining_date = get_col(row_dict, 'Joining Date', 'joining_date', 'JoiningDate', 'JOINING DATE')

    if not full_name:
        errors.append("Full Name is required")
    if not username:
        errors.append("Username is required")
    if not password:
        errors.append("Password is required")
    if not phone:
        errors.append("Phone No is required")
    if not joining_date:
        errors.append("Joining Date is required")

    if username:
        try:
            c = conn.cursor()
            c.execute("SELECT id FROM users WHERE username=%s", (username,))
            if c.fetchone():
                errors.append(f"Username '{username}' already exists")
        except Exception:
            pass

    return errors


def validate_student_row(row_dict, conn):
    errors = []

    full_name = get_col(row_dict, 'Full Name', 'full_name', 'FullName', 'FULL NAME')
    username  = get_col(row_dict, 'Username',  'username',  'USERNAME')
    password  = get_col(row_dict, 'Password',  'password',  'PASSWORD')
    phone     = get_col(row_dict, 'Phone',     'phone',     'PHONE',    'Phone No', 'phone_no')
    class_id_val = get_col(row_dict, 'Class ID', 'class_id', 'ClassID', 'CLASS ID')

    if not full_name:
        errors.append("Full Name is required")
    if not username:
        errors.append("Username is required")
    if not password:
        errors.append("Password is required")
    if not phone:
        errors.append("Phone No is required")
    if not class_id_val:
        errors.append("Class ID is required")
    elif class_id_val:
        try:
            c = conn.cursor()
            c.execute("SELECT id FROM classes WHERE id=%s", (int(float(class_id_val)),))
            if not c.fetchone():
                errors.append(f"Class ID '{class_id_val}' does not exist")
        except Exception:
            errors.append("Class ID must be a valid number")

    if username:
        try:
            c = conn.cursor()
            c.execute("SELECT id FROM users WHERE username=%s", (username,))
            if c.fetchone():
                errors.append(f"Username '{username}' already exists")
        except Exception:
            pass

    return errors

def generate_excel_report(records, upload_type):
    """Generate validation report as Excel bytes for download."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = f"{upload_type}_validation"

    skip_cols = {'id', 'uploaded_by', 'uploaded_date'}
    if not records:
        ws['A1'] = "No records found"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = [k for k in records[0].keys() if k not in skip_cols]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h.upper())
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    for ri, rec in enumerate(records, 2):
        is_valid = rec.get('validation_status') == 'VALID'
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=ri, column=ci, value=rec.get(h))
            if not is_valid:
                cell.fill = PatternFill("solid", start_color="FCE4D6")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ========== AUTH ROUTES ==========

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'GET' and 'user_id' in session:
        session.clear()

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not username or not password:
            flash('Username aur password dono zaruri hain', 'error')
            return render_template('login.html')

        conn = None
        try:
            conn = get_db()
            c = conn.cursor()
            hashed = hash_password(password)

            c.execute(
                "SELECT id, role, full_name, school_id FROM users WHERE username=%s AND password=%s AND is_active=TRUE",
                (username, hashed)
            )
            user = fetchone_dict(c)

            if user:
                session.clear()
                session['user_id'] = user['id']
                session['role'] = user['role']
                session['full_name'] = user['full_name']
                session['school_id'] = user['school_id']

                if user['role'] == 'student':
                    return redirect(url_for('my_result'))
                if user['role'] == 'parent':
                    return redirect(url_for('select_child'))
                if user['role'] in ('admin', 'school_admin'):
                    return redirect(url_for('select_school'))
                return redirect(url_for('dashboard'))

            flash('Username ya password galat hai', 'error')
        except Exception as e:
            print(f"Login error: {e}")
            flash('Technical issue. Please try again.', 'error')
        finally:
            if conn:
                conn.close()

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/logout_force')
def logout_force():
    session.clear()
    return redirect(url_for('login'))


# ========== SELECT SCHOOL (Role Based) ==========
@app.route('/select_school', methods=['GET', 'POST'])
@login_required
def select_school():
    if session['role'] == 'student':
        return redirect(url_for('my_result'))
    if session['role'] == 'parent':
        return redirect(url_for('select_child'))

    conn = get_db()
    c = conn.cursor()

    if session['role'] == 'admin':
        c.execute("SELECT id, name, address, logo FROM schools ORDER BY name")
        schools = fetchall_dict(c)
    elif session['role'] == 'school_admin':
        c.execute("SELECT id, name, address, logo FROM schools WHERE id=%s", (session['school_id'],))
        schools = fetchall_dict(c)
        if len(schools) == 1:
            session['active_school_id'] = schools[0]['id']
            conn.close()
            return redirect(url_for('dashboard'))
    else:  # Teacher
        c.execute("SELECT id, name, address, logo FROM schools WHERE id=%s", (session['school_id'],))
        schools = fetchall_dict(c)
        if len(schools) == 1:
            session['active_school_id'] = schools[0]['id']
            conn.close()
            return redirect(url_for('dashboard'))

    if request.method == 'POST':
        chosen = request.form.get('school_id')
        if chosen:
            session['active_school_id'] = int(chosen)
            conn.close()
            return redirect(url_for('dashboard'))
        flash('Please select a school', 'error')

    conn.close()
    return render_template('select_school.html', schools=schools, role=session['role'])


# ========== DASHBOARD (Role Based) ==========
@app.route('/dashboard')
@login_required
def dashboard():
    role = session['role']

    if role == 'student':
        return redirect(url_for('my_result'))

    conn = get_db()
    c = conn.cursor()

    if role == 'admin':
        school_id = session.get('active_school_id', session.get('school_id'))
        stats = {}

        c.execute("SELECT COUNT(*) FROM schools")
        stats['schools'] = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM students WHERE school_id=%s", (school_id,))
        stats['students'] = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM teachers WHERE school_id=%s", (school_id,))
        stats['teachers'] = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM classes WHERE school_id=%s", (school_id,))
        stats['classes'] = c.fetchone()[0]

        c.execute("""
            SELECT
                c.id,
                c.class_name,
                c.section,
                COUNT(DISTINCT s.id) AS total_students,
                SUM(CASE WHEN sp.avg_pct >= 40 THEN 1 ELSE 0 END) AS passed,
                SUM(CASE WHEN sp.avg_pct < 40 OR sp.avg_pct IS NULL THEN 1 ELSE 0 END) AS failed
            FROM classes c
            LEFT JOIN students s ON s.class_id = c.id
            LEFT JOIN (
                SELECT m2.student_id,
                       AVG(m2.obtained_marks * 100.0 / sub2.total_marks) AS avg_pct
                FROM marks m2
                JOIN subjects sub2 ON m2.subject_id = sub2.id
                GROUP BY m2.student_id
            ) AS sp ON sp.student_id = s.id
            WHERE c.school_id = %s
            GROUP BY c.id, c.class_name, c.section
            ORDER BY c.class_name
        """, (school_id,))
        results = fetchall_dict(c)

        c.execute("SELECT * FROM schools WHERE id=%s", (school_id,))
        school = fetchone_dict(c)

    elif role == 'school_admin':
        school_id = session.get('school_id')
        session['active_school_id'] = school_id
        stats = {}

        c.execute("SELECT COUNT(*) FROM students WHERE school_id=%s", (school_id,))
        stats['students'] = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM teachers WHERE school_id=%s", (school_id,))
        stats['teachers'] = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM classes WHERE school_id=%s", (school_id,))
        stats['classes'] = c.fetchone()[0]
        stats['schools'] = 1

        c.execute("""
            SELECT
                c.id,
                c.class_name,
                c.section,
                COUNT(DISTINCT s.id) AS total_students,
                SUM(CASE WHEN sp.avg_pct >= 40 THEN 1 ELSE 0 END) AS passed,
                SUM(CASE WHEN sp.avg_pct < 40 OR sp.avg_pct IS NULL THEN 1 ELSE 0 END) AS failed
            FROM classes c
            LEFT JOIN students s ON s.class_id = c.id
            LEFT JOIN (
                SELECT m2.student_id,
                       AVG(m2.obtained_marks * 100.0 / sub2.total_marks) AS avg_pct
                FROM marks m2
                JOIN subjects sub2 ON m2.subject_id = sub2.id
                GROUP BY m2.student_id
            ) AS sp ON sp.student_id = s.id
            WHERE c.school_id = %s
            GROUP BY c.id, c.class_name, c.section
            ORDER BY c.class_name
        """, (school_id,))
        results = fetchall_dict(c)

        c.execute("SELECT * FROM schools WHERE id=%s", (school_id,))
        school = fetchone_dict(c)

    else:  # Teacher
        school_id = session.get('school_id')
        session['active_school_id'] = school_id

        stats = {'students': 0, 'teachers': 1, 'classes': 0, 'schools': 1}
        results = []
        school = None

        c.execute("SELECT id FROM teachers WHERE user_id=%s", (session['user_id'],))
        teacher_row = c.fetchone()

        if teacher_row:
            teacher_id = teacher_row[0]

            c.execute("""
                SELECT COUNT(DISTINCT s.id) as students
                FROM students s
                JOIN classes c ON s.class_id = c.id
                JOIN teacher_classes tc ON tc.class_id = c.id
                WHERE tc.teacher_id = %s
            """, (teacher_id,))
            stats['students'] = c.fetchone()[0] or 0

            c.execute("SELECT COUNT(DISTINCT class_id) FROM teacher_classes WHERE teacher_id=%s", (teacher_id,))
            stats['classes'] = c.fetchone()[0] or 0
            stats['teachers'] = 1
            stats['schools'] = 1

            c.execute("""
                SELECT
                    c.id,
                    c.class_name,
                    c.section,
                    COUNT(DISTINCT s.id) AS total_students,
                    SUM(CASE WHEN sp.avg_pct >= 40 THEN 1 ELSE 0 END) AS passed,
                    SUM(CASE WHEN sp.avg_pct < 40 OR sp.avg_pct IS NULL THEN 1 ELSE 0 END) AS failed
                FROM teacher_classes tc
                JOIN classes c ON tc.class_id = c.id
                LEFT JOIN students s ON s.class_id = c.id
                LEFT JOIN (
                    SELECT m2.student_id,
                           AVG(m2.obtained_marks * 100.0 / sub2.total_marks) AS avg_pct
                    FROM marks m2
                    JOIN subjects sub2 ON m2.subject_id = sub2.id
                    GROUP BY m2.student_id
                ) AS sp ON sp.student_id = s.id
                WHERE tc.teacher_id = %s
                GROUP BY c.id, c.class_name, c.section
                ORDER BY c.class_name
            """, (teacher_id,))
            results = fetchall_dict(c)

        if school_id:
            c.execute("SELECT * FROM schools WHERE id=%s", (school_id,))
            school = fetchone_dict(c)

    conn.close()
    return render_template('dashboard.html', stats=stats, results=results, school=school, role=role)
# ========== SCHOOLS (Only Admin) ==========
@app.route('/schools')
@login_required
@admin_required
def schools():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM schools ORDER BY name")
    schools_list = fetchall_dict(c)
    conn.close()
    return render_template('schools.html', schools=schools_list)

@app.route('/schools/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_school():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        address = request.form.get('address', '')
        phone = request.form.get('phone', '')
        email = request.form.get('email', '')
        latitude = request.form.get('latitude', '')
        longitude = request.form.get('longitude', '')

        if not name:
            flash('School ka naam zaruri hai', 'error')
            return render_template('school_form.html')

        logo = None
        if 'logo' in request.files:
            file = request.files['logo']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                logo = filename

        conn = get_db()
        c = conn.cursor()
        c.execute(
            """INSERT INTO schools 
               (name, address, phone, email, logo, latitude, longitude)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (name, address, phone, email, logo,
             latitude if latitude else None,
             longitude if longitude else None)
        )
        conn.commit()
        conn.close()
        flash('School successfully add ho gaya!', 'success')
        return redirect(url_for('schools'))

    return render_template('school_form.html')


# ========== TEACHERS (Admin + Teacher) ==========
@app.route('/teachers')
@login_required
@teacher_required
def teachers():
    conn = get_db()
    c = conn.cursor()

    if session['role'] in ('admin', 'school_admin'):
        school_id = session.get('active_school_id') or session.get('school_id')
        c.execute("""
            SELECT t.*, s.name AS school_name
            FROM teachers t
            JOIN schools s ON t.school_id = s.id
            WHERE t.school_id = %s
            ORDER BY t.full_name
        """, (school_id,))
    else:  # Teacher
        school_id = session.get('school_id')
        c.execute("""
            SELECT t.*, s.name AS school_name
            FROM teachers t
            JOIN schools s ON t.school_id = s.id
            WHERE t.school_id = %s
            ORDER BY t.full_name
        """, (school_id,))

    teachers_list = fetchall_dict(c)
    conn.close()
    return render_template('teachers.html', teachers=teachers_list)


@app.route('/teachers/add', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def add_teacher():
    school_id = session.get('active_school_id', session.get('school_id'))
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, name FROM schools ORDER BY name")
    schools_list = fetchall_dict(c)

    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        subject_spec = request.form.get('subject_specialization', '')
        qualification = request.form.get('qualification', '')
        joining_date = request.form.get('joining_date', '')
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not full_name or not username or not password:
            flash('Naam, username aur password zaruri hain', 'error')
            conn.close()
            return render_template('teacher_form.html', schools=schools_list)

        try:
            c.execute(
                """INSERT INTO users
                   (school_id, username, password, role, full_name, email, phone)
                   VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (school_id, username, hash_password(password), 'teacher',
                 full_name, email, phone)
            )
            row = c.fetchone()
            conn.commit()
            user_id = int(row[0])
        except psycopg2.IntegrityError:
            conn.rollback()
            flash('Yeh username pehle se maujood hai!', 'error')
            conn.close()
            return render_template('teacher_form.html', schools=schools_list)

        teacher_code = f"TCH-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"
        c.execute(
            """INSERT INTO teachers
               (user_id, school_id, teacher_code, full_name, email, phone,
                subject_specialization, qualification, joining_date)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (user_id, school_id, teacher_code, full_name, email, phone,
             subject_spec, qualification, joining_date or None)
        )
        conn.commit()
        conn.close()
        flash('Teacher successfully register ho gaya!', 'success')
        return redirect(url_for('teachers'))

    conn.close()
    return render_template('teacher_form.html', schools=schools_list)


# ========== STUDENTS (Admin + Teacher) ==========
@app.route('/students')
@login_required
@teacher_required
def students():
    conn = get_db()
    c = conn.cursor()

    if session['role'] in ('admin', 'school_admin'):
        school_id = session.get('active_school_id') or session.get('school_id')
        c.execute("""
            SELECT st.*, c.class_name, c.section, s.name AS school_name
            FROM students st
            LEFT JOIN classes c ON st.class_id = c.id
            JOIN schools s ON st.school_id = s.id
            WHERE st.school_id = %s
            ORDER BY st.full_name
        """, (school_id,))
    else:  # Teacher
        c.execute("SELECT id FROM teachers WHERE user_id=%s", (session['user_id'],))
        teacher = c.fetchone()
        if teacher:
            c.execute("""
                SELECT DISTINCT st.*, c.class_name, c.section, s.name AS school_name
                FROM students st
                LEFT JOIN classes c ON st.class_id = c.id
                JOIN schools s ON st.school_id = s.id
                JOIN teacher_classes tc ON tc.class_id = c.id
                WHERE tc.teacher_id = %s
                ORDER BY st.full_name
            """, (teacher[0],))
        else:
            students_list = []
            conn.close()
            return render_template('students.html', students=[])

    students_list = fetchall_dict(c)
    conn.close()
    return render_template('students.html', students=students_list)


@app.route('/students/add', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def add_student():
    school_id = session.get('active_school_id', session.get('school_id'))
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT id, class_name, section FROM classes WHERE school_id=%s ORDER BY class_name", (school_id,))
    classes_list = fetchall_dict(c)

    c.execute("SELECT id, full_name, student_code FROM students WHERE school_id=%s ORDER BY full_name", (school_id,))
    all_students = fetchall_dict(c)
    conn.close()

    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        father_name = request.form.get('father_name', '')
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        dob = request.form.get('date_of_birth', '') or None
        gender = request.form.get('gender', '')
        class_id = request.form.get('class_id', '') or None
        joining_date = request.form.get('joining_date', '') or None
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        address_line1 = request.form.get('address_line1', '')
        address_line2 = request.form.get('address_line2', '')
        city = request.form.get('city', '')
        state = request.form.get('state', '')
        postal_code = request.form.get('postal_code', '')
        medical_details = request.form.get('medical_details', '')

        if not full_name or not username or not password:
            flash('Naam, username aur password zaruri hain', 'error')
            return render_template('student_form.html',
                                   classes=classes_list,
                                   all_students=all_students,
                                   now=datetime.datetime.now())

        conn = get_db()
        c = conn.cursor()

        try:
            c.execute("""
                INSERT INTO users 
                (school_id, username, password, role, full_name, email, phone)
                VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
            """, (school_id, username, hash_password(password), 'student',
                  full_name, email, phone))

            row = c.fetchone()
            user_id = int(row[0])

            profile_pic_filename = None
            if 'profile_pic' in request.files:
                file = request.files['profile_pic']
                if file and file.filename and allowed_file(file.filename):
                    profile_pic_filename = secure_filename(
                        f"student_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
                    )
                    os.makedirs('static/uploads/students', exist_ok=True)
                    file.save(os.path.join('static/uploads/students', profile_pic_filename))

            student_code = f"STD-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"

            c.execute("""
                INSERT INTO students 
                (user_id, school_id, student_code, full_name, father_name, email,
                 phone, date_of_birth, gender, class_id, joining_date,
                 address_line1, address_line2, city, state, postal_code,
                 medical_details, profile_pic, created_by, created_date)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW()) RETURNING id
            """, (user_id, school_id, student_code, full_name, father_name, email,
                  phone, dob, gender, class_id, joining_date,
                  address_line1, address_line2, city, state, postal_code,
                  medical_details, profile_pic_filename, session['user_id']))

            student_id = c.fetchone()[0]

            parent_names = request.form.getlist('parent_name[]')
            parent_relations = request.form.getlist('parent_relation[]')
            parent_occupations = request.form.getlist('parent_occupation[]')
            parent_phones = request.form.getlist('parent_phone[]')
            parent_emails = request.form.getlist('parent_email[]')

            for i in range(len(parent_names)):
                if parent_names[i].strip():
                    c.execute("""
                        INSERT INTO student_parents 
                        (student_id, parent_name, relation, occupation, phone, email, created_by)
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """, (student_id, parent_names[i], parent_relations[i] if i < len(parent_relations) else 'Guardian',
                          parent_occupations[i] if i < len(parent_occupations) else '',
                          parent_phones[i] if i < len(parent_phones) else '',
                          parent_emails[i] if i < len(parent_emails) else '',
                          session['user_id']))

            sibling_names = request.form.getlist('sibling_name[]')
            sibling_ages = request.form.getlist('sibling_age[]')
            sibling_classes = request.form.getlist('sibling_class[]')
            sibling_schools = request.form.getlist('sibling_school_name[]')
            sibling_student_ids = request.form.getlist('sibling_student_id[]')

            for i in range(len(sibling_names)):
                if sibling_names[i].strip():
                    sib_student_id = sibling_student_ids[i] if sibling_student_ids and sibling_student_ids[i] else None
                    if sib_student_id == '' or sib_student_id == 'None':
                        sib_student_id = None

                    c.execute("""
                        INSERT INTO student_siblings 
                        (student_id, sibling_name, age, class, school_name, sibling_student_id, created_by)
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """, (student_id, sibling_names[i],
                          sibling_ages[i] if i < len(sibling_ages) else None,
                          sibling_classes[i] if i < len(sibling_classes) else '',
                          sibling_schools[i] if i < len(sibling_schools) else '',
                          sib_student_id, session['user_id']))

            conn.commit()
            flash('Student successfully register ho gaya!', 'success')
            return redirect(url_for('students'))

        except psycopg2.IntegrityError as e:
            conn.rollback()
            flash(f'Error: {str(e)}', 'error')
            print(f"IntegrityError: {e}")
        except Exception as e:
            conn.rollback()
            flash(f'Error: {str(e)}', 'error')
            print(f"Exception: {e}")
        finally:
            conn.close()

    return render_template('student_form.html',
                           classes=classes_list,
                           all_students=all_students,
                           now=datetime.datetime.now())

# ========== CLASSES (Admin + Teacher) ==========
@app.route('/classes')
@login_required
@teacher_required
def classes():
    conn = get_db()
    c = conn.cursor()

    if session['role'] in ('admin', 'school_admin'):
        school_id = session.get('active_school_id') or session.get('school_id')
        c.execute("""
            SELECT 
                c.*,
                s.name AS school_name,
                (SELECT COUNT(*) FROM students st WHERE st.class_id = c.id) AS student_count
            FROM classes c
            JOIN schools s ON c.school_id = s.id
            WHERE c.school_id = %s
            ORDER BY c.class_name
        """, (school_id,))
    else:  # Teacher
        c.execute("SELECT id FROM teachers WHERE user_id=%s", (session['user_id'],))
        teacher = c.fetchone()
        if teacher:
            c.execute("""
                SELECT 
                    c.*,
                    s.name AS school_name,
                    (SELECT COUNT(*) FROM students st WHERE st.class_id = c.id) AS student_count
                FROM classes c
                JOIN schools s ON c.school_id = s.id
                JOIN teacher_classes tc ON tc.class_id = c.id
                WHERE tc.teacher_id = %s
                ORDER BY c.class_name
            """, (teacher[0],))
        else:
            classes_list = []
            conn.close()
            return render_template('classes.html', classes=[])

    classes_list = fetchall_dict(c)
    conn.close()
    return render_template('classes.html', classes=classes_list)


# ========== SUBJECTS (Admin + Teacher) ==========
@app.route('/subjects', methods=['GET', 'POST'])
@login_required
@teacher_required
def subjects():
    conn = get_db()
    c = conn.cursor()

    if session['role'] in ('admin', 'school_admin'):
        school_id = session.get('active_school_id') or session.get('school_id')
        c.execute("SELECT * FROM classes WHERE school_id=%s ORDER BY class_name", (school_id,))
        classes_list = fetchall_dict(c)

        if request.method == 'POST':
            subject_name = request.form.get('subject_name', '').strip()
            class_id = request.form.get('class_id')
            total_marks = request.form.get('total_marks', 100)
            passing_marks = request.form.get('passing_marks', 40)

            if not subject_name or not class_id:
                flash('Subject ka naam aur class zaruri hain', 'error')
            else:
                c.execute(
                    "INSERT INTO subjects (school_id, class_id, subject_name, total_marks, passing_marks) VALUES (%s,%s,%s,%s,%s)",
                    (school_id, class_id, subject_name, total_marks, passing_marks)
                )
                conn.commit()
                flash('Subject add ho gaya!', 'success')

        c.execute("""
            SELECT sub.*, c.class_name, c.section
            FROM subjects sub
            JOIN classes c ON sub.class_id = c.id
            WHERE sub.school_id = %s
            ORDER BY c.class_name, sub.subject_name
        """, (school_id,))

    else:  # Teacher
        c.execute("SELECT id FROM teachers WHERE user_id=%s", (session['user_id'],))
        teacher = c.fetchone()
        if teacher:
            c.execute("""
                SELECT DISTINCT c.id, c.class_name, c.section
                FROM classes c
                JOIN teacher_classes tc ON tc.class_id = c.id
                WHERE tc.teacher_id = %s
                ORDER BY c.class_name
            """, (teacher[0],))
            classes_list = fetchall_dict(c)

            if request.method == 'POST':
                subject_name = request.form.get('subject_name', '').strip()
                class_id = request.form.get('class_id')
                total_marks = request.form.get('total_marks', 100)
                passing_marks = request.form.get('passing_marks', 40)

                if not subject_name or not class_id:
                    flash('Subject ka naam aur class zaruri hain', 'error')
                else:
                    c.execute(
                        "INSERT INTO subjects (school_id, class_id, subject_name, total_marks, passing_marks) VALUES (%s,%s,%s,%s,%s)",
                        (session['school_id'], class_id, subject_name, total_marks, passing_marks)
                    )
                    conn.commit()
                    flash('Subject add ho gaya!', 'success')
            c.execute("""
                SELECT sub.id, sub.subject_name, sub.total_marks, 
                       sub.passing_marks, sub.school_id, sub.class_id,
                       c.class_name, c.section
                FROM subjects sub
                JOIN classes c ON sub.class_id = c.id
                WHERE sub.school_id = %s
                  AND c.school_id = %s
                  AND sub.class_id IN (
                      SELECT tc.class_id 
                      FROM teacher_classes tc 
                      WHERE tc.teacher_id = %s
                  )
                ORDER BY c.class_name, sub.subject_name
            """, (session['school_id'], session['school_id'], teacher[0]))

        else:
            classes_list = []
            subjects_list = []

    subjects_list = fetchall_dict(c)
    conn.close()
    return render_template('subjects.html', subjects=subjects_list, classes=classes_list)


# ========== MARKS ENTRY (Admin + Teacher) ==========
@app.route('/marks', methods=['GET', 'POST'])
@login_required
@teacher_required
def marks():
    conn = get_db()
    c = conn.cursor()
    role = session['role']

    teacher = None
    classes_list = []

    if role in ('admin', 'school_admin'):
        school_id = session.get('active_school_id') or session.get('school_id')
        c.execute("SELECT * FROM classes WHERE school_id=%s ORDER BY class_name", (school_id,))
        classes_list = fetchall_dict(c)
        c.execute("SELECT * FROM teachers WHERE school_id=%s", (school_id,))
        teachers_list = fetchall_dict(c)
    else:  # Teacher
        c.execute("SELECT * FROM teachers WHERE user_id=%s", (session['user_id'],))
        teacher = fetchone_dict(c)
        if teacher:
            c.execute("""
                SELECT c.* FROM classes c
                JOIN teacher_classes tc ON tc.class_id = c.id
                WHERE tc.teacher_id = %s
                ORDER BY c.class_name
            """, (teacher['id'],))
            classes_list = fetchall_dict(c)
        teachers_list = []

    if request.method == 'POST':
        class_id = request.form.get('class_id')
        subject_id = request.form.get('subject_id')
        exam_type = request.form.get('exam_type', 'Annual')
        academic_year = request.form.get('academic_year', str(datetime.datetime.now().year))

        if role == 'teacher' and teacher:
            teacher_id = teacher['id']
            c.execute("SELECT id FROM teacher_classes WHERE teacher_id=%s AND class_id=%s", (teacher_id, class_id))
            if not c.fetchone():
                flash('Aapko is class ke marks enter karne ki permission nahi hai', 'error')
                conn.close()
                return redirect(url_for('marks'))
        else:
            teacher_id = request.form.get('teacher_id')

        student_ids = request.form.getlist('student_id[]')
        obtained_list = request.form.getlist('obtained_marks[]')

        saved = 0
        for sid, marks_val in zip(student_ids, obtained_list):
            if not marks_val.strip():
                continue
            c.execute(
                "SELECT id FROM marks WHERE student_id=%s AND subject_id=%s AND exam_type=%s AND academic_year=%s",
                (sid, subject_id, exam_type, academic_year)
            )
            existing = c.fetchone()
            if existing:
                c.execute(
                    "UPDATE marks SET obtained_marks=%s, entered_at=NOW() WHERE id=%s",
                    (marks_val, existing[0])
                )
            else:
                c.execute(
                    """INSERT INTO marks
                       (student_id, subject_id, class_id, school_id, teacher_id,
                        obtained_marks, exam_type, academic_year)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (sid, subject_id, class_id, session.get('school_id') or session.get('active_school_id'),
                     teacher_id, marks_val, exam_type, academic_year)
                )
            saved += 1

        conn.commit()
        conn.close()
        flash(f'{saved} students ke marks save ho gaye!', 'success')
        return redirect(url_for('marks'))

    conn.close()
    return render_template('marks_entry.html',
                           classes=classes_list,
                           teachers=teachers_list,
                           teacher=teacher,
                           role=role,
                           now=datetime.datetime.now())


# ========== RESULT CARD (All Roles) ==========
@app.route('/result/<int:student_id>')
@login_required
def result_card(student_id):
    conn = get_db()
    c = conn.cursor()
    role = session['role']

    if role == 'student':
        c.execute("SELECT id FROM students WHERE user_id=%s", (session['user_id'],))
        student_record = c.fetchone()
        if not student_record or student_record[0] != student_id:
            flash('Aap sirf apna result dekh sakte hain', 'error')
            conn.close()
            return redirect(url_for('dashboard'))
    elif role == 'teacher':
        c.execute("SELECT class_id FROM students WHERE id=%s", (student_id,))
        student_class = c.fetchone()
        if student_class:
            c.execute("""
                SELECT id FROM teacher_classes 
                WHERE teacher_id = (SELECT id FROM teachers WHERE user_id=%s) 
                AND class_id = %s
            """, (session['user_id'], student_class[0]))
            if not c.fetchone():
                flash('Aap sirf apne class ke students ka result dekh sakte hain', 'error')
                conn.close()
                return redirect(url_for('dashboard'))



    elif role == 'parent':
       c.execute("""
                SELECT id FROM parent_children 
                WHERE parent_user_id=%s AND student_id=%s
       """, (session['user_id'], student_id))
    if not c.fetchone():
        flash('Aap is student ka result nahi dekh sakte', 'error')
        conn.close()
        return redirect(url_for('parent_dashboard'))
    exam_type = request.args.get('exam_type', 'Annual')
    academic_year = request.args.get('academic_year', str(datetime.datetime.now().year))

    c.execute("""
        SELECT st.*, c.class_name, c.section,
               s.name AS school_name,
               s.address AS school_address,
               s.logo AS school_logo
        FROM students st
        LEFT JOIN classes c ON st.class_id = c.id
        JOIN schools s ON st.school_id = s.id
        WHERE st.id = %s
    """, (student_id,))
    student = fetchone_dict(c)

    if not student:
        flash('Student nahi mila', 'error')
        conn.close()
        return redirect(url_for('students' if role != 'student' else 'dashboard'))

    c.execute("""
        SELECT m.*, sub.subject_name, sub.total_marks, sub.passing_marks,
               t.full_name AS teacher_name
        FROM marks m
        JOIN subjects sub ON m.subject_id = sub.id
        JOIN teachers t ON m.teacher_id = t.id
        WHERE m.student_id = %s AND m.exam_type = %s AND m.academic_year = %s
        ORDER BY sub.subject_name
    """, (student_id, exam_type, academic_year))
    marks_data = fetchall_dict(c)

    total_obtained = sum(m['obtained_marks'] for m in marks_data)
    total_marks_sum = sum(m['total_marks'] for m in marks_data)
    percentage = (total_obtained / total_marks_sum * 100) if total_marks_sum > 0 else 0

    if percentage >= 90:
        grade = 'A+'
    elif percentage >= 80:
        grade = 'A'
    elif percentage >= 70:
        grade = 'B'
    elif percentage >= 60:
        grade = 'C'
    elif percentage >= 50:
        grade = 'D'
    else:
        grade = 'F'

    status = 'PASS' if (
            percentage >= 40 and all(m['obtained_marks'] >= m['passing_marks'] for m in marks_data)) else 'FAIL'

    class_teacher = None
    if student.get('class_id'):
        c.execute("""
            SELECT t.full_name
            FROM teacher_classes tc
            JOIN teachers t ON tc.teacher_id = t.id
            WHERE tc.class_id = %s AND tc.is_primary = TRUE
            LIMIT 1
        """, (student['class_id'],))
        class_teacher = fetchone_dict(c)

    conn.close()
    return render_template('result_card.html',
                           student=student,
                           marks=marks_data,
                           total_obtained=total_obtained,
                           total_marks=total_marks_sum,
                           percentage=round(percentage, 2),
                           grade=grade,
                           status=status,
                           class_teacher=class_teacher,
                           exam_type=exam_type,
                           academic_year=academic_year)


# ========== STUDENT MY RESULT ==========
@app.route('/my_result')
@login_required
def my_result():
    if session['role'] != 'student':
        return redirect(url_for('dashboard'))

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM students WHERE user_id=%s", (session['user_id'],))
    student = fetchone_dict(c)
    conn.close()

    if student:
        return redirect(url_for('result_card', student_id=student['id']))

    flash('Aapka student record nahi mila', 'error')
    return redirect(url_for('dashboard'))


# ========== OTHER ROUTES (Admin Only) ==========
@app.route('/users')
@login_required
@school_admin_only_required
def users():
    conn = get_db()
    c = conn.cursor()
    school_id = session.get('active_school_id') or session.get('school_id')
    c.execute("""
        SELECT u.*, s.name AS school_name
        FROM users u
        LEFT JOIN schools s ON u.school_id = s.id
        WHERE u.school_id = %s
        ORDER BY u.role, u.full_name
    """, (school_id,))
    users_list = fetchall_dict(c)
    conn.close()
    return render_template('users.html', users=users_list)


@app.route('/users/toggle/<int:user_id>')
@login_required
@school_admin_or_super_admin_required
def toggle_user(user_id):
    conn = get_db()
    c = conn.cursor()
    school_id = session.get('active_school_id') or session.get('school_id')
    c.execute("SELECT is_active, school_id FROM users WHERE id=%s", (user_id,))
    row = c.fetchone()
    if not row:
        flash('User nahi mila', 'error')
        conn.close()
        return redirect(url_for('users'))

    if session.get('role') == 'school_admin' and row[1] != school_id:
        flash('Aap sirf apne school ke users ko manage kar sakte hain', 'error')
        conn.close()
        return redirect(url_for('users'))

    new_status = not row[0]
    c.execute("UPDATE users SET is_active=%s WHERE id=%s", (new_status, user_id))
    conn.commit()
    conn.close()
    flash('User status update ho gaya!', 'success')
    return redirect(url_for('users'))


@app.route('/assignments', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def assignments():
    school_id = session.get('active_school_id', session.get('school_id'))
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM teachers WHERE school_id=%s ORDER BY full_name", (school_id,))
    teachers_list = fetchall_dict(c)

    c.execute("SELECT * FROM classes WHERE school_id=%s ORDER BY class_name", (school_id,))
    classes_list = fetchall_dict(c)

    if request.method == 'POST':
        teacher_id = request.form.get('teacher_id')
        class_ids = request.form.getlist('class_ids')
        is_primary = bool(int(request.form.get('is_primary', 0)))

        if not teacher_id or not class_ids:
            flash('Teacher aur class dono select karein', 'error')
        else:
            for class_id in class_ids:
                c.execute(
                    "SELECT id FROM teacher_classes WHERE teacher_id=%s AND class_id=%s",
                    (teacher_id, class_id)
                )
                if not c.fetchone():
                    c.execute(
                        "INSERT INTO teacher_classes (teacher_id, class_id, school_id, is_primary) VALUES (%s,%s,%s,%s)",
                        (teacher_id, class_id, school_id, is_primary)
                    )
            conn.commit()
            flash('Teacher class ko assign ho gaya!', 'success')

    c.execute("""
        SELECT tc.*, t.full_name AS teacher_name, c.class_name, c.section
        FROM teacher_classes tc
        JOIN teachers t ON tc.teacher_id = t.id
        JOIN classes c ON tc.class_id = c.id
        WHERE tc.school_id = %s
        ORDER BY t.full_name
    """, (school_id,))
    assignments_list = fetchall_dict(c)

    conn.close()
    return render_template('assignments.html',
                           teachers=teachers_list,
                           classes=classes_list,
                           assignments=assignments_list)


# ========== AJAX ENDPOINTS ==========
@app.route('/get_students/<int:class_id>')
@login_required
@teacher_required
def get_students(class_id):
    conn = get_db()
    c = conn.cursor()

    if session['role'] == 'teacher':
        c.execute("""
            SELECT id FROM teacher_classes 
            WHERE teacher_id = (SELECT id FROM teachers WHERE user_id=%s) 
            AND class_id = %s
        """, (session['user_id'], class_id))
        if not c.fetchone():
            return jsonify([])

    c.execute(
        "SELECT id, full_name, student_code FROM students WHERE class_id=%s ORDER BY full_name",
        (class_id,)
    )
    students = fetchall_dict(c)
    conn.close()
    return jsonify(students)


@app.route('/get_subjects/<int:class_id>')
@login_required
@teacher_required
def get_subjects(class_id):
    conn = get_db()
    c = conn.cursor()

    school_id = session.get('active_school_id') or session.get('school_id')
    if session['role'] == 'teacher':
        c.execute("""
            SELECT sub.* FROM subjects sub
            JOIN classes c ON sub.class_id = c.id
            JOIN teacher_classes tc ON tc.class_id = c.id
            WHERE sub.class_id = %s 
              AND sub.school_id = %s
              AND tc.teacher_id = (SELECT id FROM teachers WHERE user_id = %s)
            ORDER BY sub.subject_name
        """, (class_id, school_id, session['user_id']))
    else:
        c.execute(
            "SELECT * FROM subjects WHERE class_id=%s AND school_id=%s ORDER BY subject_name",
            (class_id, school_id)
        )

    subjects_list = fetchall_dict(c)
    conn.close()
    return jsonify(subjects_list)


# ========== TEACHER SUBJECTS ASSIGNMENT ==========
@app.route('/teacher_subjects', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def teacher_subjects():
    school_id = session.get('active_school_id', session.get('school_id'))
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM teachers WHERE school_id=%s ORDER BY full_name", (school_id,))
    teachers_list = fetchall_dict(c)

    c.execute("""
        SELECT sub.*, c.class_name, c.section
        FROM subjects sub
        JOIN classes c ON sub.class_id = c.id
        WHERE sub.school_id = %s
        ORDER BY c.class_name, sub.subject_name
    """, (school_id,))
    subjects_list = fetchall_dict(c)

    if request.method == 'POST':
        teacher_id = request.form.get('teacher_id')
        subject_ids = request.form.getlist('subject_ids')

        if not teacher_id or not subject_ids:
            flash('Teacher aur subject dono select karein', 'error')
        else:
            for subject_id in subject_ids:
                c.execute(
                    "SELECT id FROM teacher_subjects WHERE teacher_id=%s AND subject_id=%s",
                    (teacher_id, subject_id)
                )
                if not c.fetchone():
                    c.execute(
                        "INSERT INTO teacher_subjects (teacher_id, subject_id, school_id) VALUES (%s,%s,%s)",
                        (teacher_id, subject_id, school_id)
                    )
            conn.commit()
            flash('Subjects assign ho gaye!', 'success')

    c.execute("""
        SELECT ts.*, t.full_name AS teacher_name, 
               sub.subject_name, c.class_name, c.section
        FROM teacher_subjects ts
        JOIN teachers t ON ts.teacher_id = t.id
        JOIN subjects sub ON ts.subject_id = sub.id
        JOIN classes c ON sub.class_id = c.id
        WHERE ts.school_id = %s
        ORDER BY t.full_name
    """, (school_id,))
    assignments_list = fetchall_dict(c)

    conn.close()
    return render_template('teacher_subjects.html',
                           teachers=teachers_list,
                           subjects=subjects_list,
                           assignments=assignments_list)


@app.route('/classes/add', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def add_class():
    school_id = session.get('active_school_id', session.get('school_id'))
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT id, name FROM schools ORDER BY name")
    schools_list = fetchall_dict(c)

    if request.method == 'POST':
        school_id = request.form.get('school_id')
        class_name = request.form.get('class_name', '').strip()
        section = request.form.get('section', '')
        academic_year = request.form.get('academic_year', '')

        if not class_name:
            flash('Class ka naam zaruri hai', 'error')
            conn.close()
            return render_template('class_form.html', schools=schools_list)

        c.execute(
            "INSERT INTO classes (school_id, class_name, section, academic_year) VALUES (%s,%s,%s,%s)",
            (school_id, class_name, section, academic_year)
        )
        conn.commit()
        conn.close()
        flash('Class successfully add ho gayi!', 'success')
        return redirect(url_for('classes'))

    conn.close()
    return render_template('class_form.html', schools=schools_list, now=datetime.datetime.now())


@app.route('/classes/<int:class_id>/students')
@login_required
@teacher_required
def class_students(class_id):
    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT c.*, s.name AS school_name 
        FROM classes c
        JOIN schools s ON c.school_id = s.id
        WHERE c.id = %s
    """, (class_id,))
    class_info = fetchone_dict(c)

    if not class_info:
        flash('Class nahi mili', 'error')
        conn.close()
        return redirect(url_for('classes'))

    if session['role'] == 'teacher':
        c.execute("""
            SELECT id FROM teacher_classes 
            WHERE teacher_id = (SELECT id FROM teachers WHERE user_id=%s) 
            AND class_id = %s
        """, (session['user_id'], class_id))
        if not c.fetchone():
            flash('Aapko is class ka access nahi hai', 'error')
            conn.close()
            return redirect(url_for('classes'))

    c.execute("""
        SELECT st.*, c.class_name, c.section
        FROM students st
        LEFT JOIN classes c ON st.class_id = c.id
        WHERE st.class_id = %s
        ORDER BY st.full_name
    """, (class_id,))
    students_list = fetchall_dict(c)

    conn.close()
    return render_template('class_students.html',
                           students=students_list,
                           class_info=class_info)


# ========== STUDENT ENHANCED ROUTES ==========
@app.route('/students/<int:student_id>/edit', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def edit_student(student_id):
    conn = get_db()
    c = conn.cursor()

    school_id = session.get('active_school_id') or session.get('school_id')

    c.execute("SELECT school_id FROM students WHERE id=%s", (student_id,))
    st_row = c.fetchone()
    if st_row and session.get('role') == 'school_admin' and st_row[0] != school_id:
        flash('Aap sirf apne school ke students edit kar sakte hain', 'error')
        conn.close()
        return redirect(url_for('students'))

    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        father_name = request.form.get('father_name', '')
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        dob = request.form.get('date_of_birth', '') or None
        gender = request.form.get('gender', '')
        class_id = request.form.get('class_id', '') or None
        address_line1 = request.form.get('address_line1', '')
        address_line2 = request.form.get('address_line2', '')
        city = request.form.get('city', '')
        state = request.form.get('state', '')
        postal_code = request.form.get('postal_code', '')
        medical_details = request.form.get('medical_details', '')

        profile_pic = None
        if 'profile_pic' in request.files:
            file = request.files['profile_pic']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(f"student_{student_id}_{file.filename}")
                os.makedirs('static/uploads/students', exist_ok=True)
                file.save(os.path.join('static/uploads/students', filename))
                profile_pic = filename

        if profile_pic:
            c.execute("""
                UPDATE students 
                SET full_name=%s, father_name=%s, email=%s, phone=%s,
                    date_of_birth=%s, gender=%s, class_id=%s,
                    address_line1=%s, address_line2=%s, city=%s, state=%s,
                    postal_code=%s, medical_details=%s, profile_pic=%s,
                    updated_by=%s, updated_date=NOW()
                WHERE id=%s
            """, (full_name, father_name, email, phone, dob, gender, class_id,
                  address_line1, address_line2, city, state, postal_code,
                  medical_details, profile_pic, session['user_id'], student_id))
        else:
            c.execute("""
                UPDATE students 
                SET full_name=%s, father_name=%s, email=%s, phone=%s,
                    date_of_birth=%s, gender=%s, class_id=%s,
                    address_line1=%s, address_line2=%s, city=%s, state=%s,
                    postal_code=%s, medical_details=%s,
                    updated_by=%s, updated_date=NOW()
                WHERE id=%s
            """, (full_name, father_name, email, phone, dob, gender, class_id,
                  address_line1, address_line2, city, state, postal_code,
                  medical_details, session['user_id'], student_id))

        parent_names = request.form.getlist('parent_name[]')
        parent_relations = request.form.getlist('parent_relation[]')
        parent_occupations = request.form.getlist('parent_occupation[]')
        parent_phones = request.form.getlist('parent_phone[]')
        parent_emails = request.form.getlist('parent_email[]')

        c.execute("DELETE FROM student_parents WHERE student_id=%s", (student_id,))

        for i in range(len(parent_names)):
            if parent_names[i].strip():
                c.execute("""
                    INSERT INTO student_parents 
                    (student_id, parent_name, relation, occupation, phone, email, created_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """, (student_id, parent_names[i], parent_relations[i],
                      parent_occupations[i], parent_phones[i], parent_emails[i],
                      session['user_id']))

        sibling_names = request.form.getlist('sibling_name[]')
        sibling_ages = request.form.getlist('sibling_age[]')
        sibling_classes = request.form.getlist('sibling_class[]')
        sibling_schools = request.form.getlist('sibling_school_name[]')
        sibling_student_ids = request.form.getlist('sibling_student_id[]')

        c.execute("DELETE FROM student_siblings WHERE student_id=%s", (student_id,))

        for i in range(len(sibling_names)):
            if sibling_names[i].strip():
                sib_student_id = sibling_student_ids[i] if sibling_student_ids[i] else None
                c.execute("""
                    INSERT INTO student_siblings 
                    (student_id, sibling_name, age, class, school_name, 
                     sibling_student_id, created_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """, (student_id, sibling_names[i], sibling_ages[i],
                      sibling_classes[i], sibling_schools[i],
                      sib_student_id, session['user_id']))

        conn.commit()
        conn.close()
        flash('Student details update ho gaye!', 'success')
        return redirect(url_for('students'))

    c.execute("""
        SELECT st.*, u.username 
        FROM students st 
        JOIN users u ON st.user_id = u.id 
        WHERE st.id = %s
    """, (student_id,))
    student = fetchone_dict(c)

    if not student:
        flash('Student nahi mila', 'error')
        conn.close()
        return redirect(url_for('students'))

    c.execute("SELECT * FROM student_parents WHERE student_id=%s", (student_id,))
    parents = fetchall_dict(c)

    c.execute("""
        SELECT ss.*, s.full_name as linked_name 
        FROM student_siblings ss
        LEFT JOIN students s ON ss.sibling_student_id = s.id
        WHERE ss.student_id = %s
    """, (student_id,))
    siblings = fetchall_dict(c)

    c.execute("SELECT id, class_name, section FROM classes WHERE school_id=%s ORDER BY class_name", (school_id,))
    classes_list = fetchall_dict(c)

    c.execute("SELECT id, full_name, student_code FROM students WHERE school_id=%s ORDER BY full_name", (school_id,))
    all_students = fetchall_dict(c)

    conn.close()
    return render_template('student_edit.html',
                           student=student,
                           parents=parents,
                           siblings=siblings,
                           classes=classes_list,
                           all_students=all_students)


# ========== TEACHER ENHANCED ROUTES ==========
@app.route('/teachers/<int:teacher_id>/edit', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def edit_teacher(teacher_id):
    conn = get_db()
    c = conn.cursor()

    school_id = session.get('active_school_id') or session.get('school_id')
    c.execute("SELECT school_id FROM teachers WHERE id=%s", (teacher_id,))
    t_row = c.fetchone()
    if t_row and session.get('role') == 'school_admin' and t_row[0] != school_id:
        flash('Aap sirf apne school ke teachers edit kar sakte hain', 'error')
        conn.close()
        return redirect(url_for('teachers'))

    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        subject_spec = request.form.get('subject_specialization', '')
        qualification = request.form.get('qualification', '')
        salary = request.form.get('salary', '') or None
        address = request.form.get('address', '')
        joining_date = request.form.get('joining_date', '') or None
        cnic = request.form.get('cnic', '')

        profile_pic = None
        if 'profile_pic' in request.files:
            file = request.files['profile_pic']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(
                    f"teacher_{teacher_id}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                os.makedirs('static/uploads/teachers', exist_ok=True)
                file.save(os.path.join('static/uploads/teachers', filename))
                profile_pic = filename

        if profile_pic:
            c.execute("""
                UPDATE teachers
                SET full_name=%s, email=%s, phone=%s, subject_specialization=%s,
                    qualification=%s, salary=%s, profile_pic=%s, address=%s,
                    joining_date=%s, cnic=%s,
                    updated_by=%s, updated_date=NOW()
                WHERE id=%s
            """, (full_name, email, phone, subject_spec, qualification,
                  salary, profile_pic, address, joining_date, cnic,
                  session['user_id'], teacher_id))
        else:
            c.execute("""
                UPDATE teachers
                SET full_name=%s, email=%s, phone=%s, subject_specialization=%s,
                    qualification=%s, salary=%s, address=%s,
                    joining_date=%s, cnic=%s,
                    updated_by=%s, updated_date=NOW()
                WHERE id=%s
            """, (full_name, email, phone, subject_spec, qualification,
                  salary, address, joining_date, cnic,
                  session['user_id'], teacher_id))

        doc_types = request.form.getlist('doc_type[]')
        doc_names = request.form.getlist('doc_name[]')

        if 'documents' in request.files:
            files = request.files.getlist('documents')
            for idx, file in enumerate(files):
                if file and file.filename:
                    ext = file.filename.rsplit('.', 1)[-1].lower()
                    allowed_docs = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx'}
                    if ext in allowed_docs:
                        cert_filename = secure_filename(
                            f"doc_{teacher_id}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{idx}_{file.filename}"
                        )
                        os.makedirs('static/uploads/teacher_docs', exist_ok=True)
                        file.save(os.path.join('static/uploads/teacher_docs', cert_filename))
                        doc_type = doc_types[idx] if idx < len(doc_types) else 'Document'
                        doc_name = doc_names[idx] if idx < len(doc_names) else file.filename
                        c.execute("""
                            INSERT INTO teacher_documents
                            (teacher_id, document_type, document_name, file_path, created_by)
                            VALUES (%s,%s,%s,%s,%s)
                        """, (teacher_id, doc_type, doc_name, cert_filename, session['user_id']))

        conn.commit()
        conn.close()
        flash('Teacher details update ho gaye!', 'success')
        return redirect(url_for('teachers'))

    c.execute("SELECT * FROM teachers WHERE id=%s", (teacher_id,))
    teacher = fetchone_dict(c)

    if not teacher:
        flash('Teacher nahi mila', 'error')
        conn.close()
        return redirect(url_for('teachers'))

    c.execute("SELECT * FROM teacher_documents WHERE teacher_id=%s ORDER BY id DESC", (teacher_id,))
    documents = fetchall_dict(c)

    conn.close()
    return render_template('teacher_edit.html', teacher=teacher, documents=documents)


@app.route('/teachers/document/delete/<int:doc_id>')
@login_required
@school_admin_only_required
def delete_teacher_document(doc_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT file_path FROM teacher_documents WHERE id=%s", (doc_id,))
    row = c.fetchone()
    if row:
        filepath = os.path.join('static/uploads/teacher_docs', row[0])
        if os.path.exists(filepath):
            os.remove(filepath)
        c.execute("DELETE FROM teacher_documents WHERE id=%s", (doc_id,))
        conn.commit()
        flash('Document delete ho gaya!', 'success')
    conn.close()
    return redirect(request.referrer or url_for('teachers'))


# ========== TEACHER ATTENDANCE (Admin only, all schools) ==========
@app.route('/attendance/teacher', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def teacher_attendance():
    school_id = session.get('active_school_id', session.get('school_id'))
    today = datetime.date.today()

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM teachers WHERE school_id=%s ORDER BY full_name", (school_id,))
    teachers = fetchall_dict(c)

    attendance_date = request.form.get('attendance_date', str(today)) \
        if request.method == 'POST' else str(today)

    existing = {}
    if attendance_date:
        c.execute("""
            SELECT teacher_id, status
            FROM teacher_attendance
            WHERE school_id=%s AND attendance_date=%s
        """, (school_id, attendance_date))
        for row in c.fetchall():
            existing[row[0]] = row[1]

    conn.close()
    return render_template('teacher_attendance.html',
                           teachers=teachers,
                           attendance_date=attendance_date,
                           existing=existing,
                           today=today)


@app.route('/attendance/teacher/save', methods=['POST'])
@login_required
@school_admin_only_required
def save_teacher_attendance():
    school_id = session.get('active_school_id', session.get('school_id'))
    attendance_date = request.form.get('attendance_date')
    marked_by = session['user_id']

    if not attendance_date:
        flash('Date zaruri hai', 'error')
        return redirect(url_for('teacher_attendance'))

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT id FROM teachers WHERE school_id=%s ORDER BY full_name", (school_id,))
    teachers = c.fetchall()

    if not teachers:
        flash('Koi teacher nahi mila', 'error')
        conn.close()
        return redirect(url_for('teacher_attendance'))

    saved_count = 0
    for teacher_row in teachers:
        teacher_id = teacher_row[0]

        status = request.form.get(f'status_{teacher_id}', 'Present')

        # Postgres: INSERT ... ON CONFLICT (upsert) instead of SQL Server MERGE
        c.execute("""
            INSERT INTO teacher_attendance (teacher_id, school_id, attendance_date, status, marked_by)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (teacher_id, attendance_date)
            DO UPDATE SET status = EXCLUDED.status, marked_by = EXCLUDED.marked_by
        """, (teacher_id, school_id, attendance_date, status, marked_by))

        saved_count += 1

    conn.commit()
    conn.close()
    flash(f'{saved_count} teachers ki attendance save ho gayi!', 'success')
    return redirect(url_for('teacher_attendance'))


# Teacher attendance history / report
@app.route('/attendance/teacher/report')
@login_required
@school_admin_only_required
def teacher_attendance_report():
    school_id = session.get('active_school_id', session.get('school_id'))
    month = request.args.get('month', datetime.date.today().strftime('%Y-%m'))

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT t.full_name, t.id as teacher_id,
               ta.attendance_date, ta.status
        FROM teachers t
        LEFT JOIN teacher_attendance ta
            ON ta.teacher_id = t.id
            AND TO_CHAR(ta.attendance_date, 'YYYY-MM') = %s
        WHERE t.school_id = %s
        ORDER BY t.full_name, ta.attendance_date
    """, (month, school_id))
    rows = fetchall_dict(c)
    conn.close()

    from collections import defaultdict
    import calendar

    year_num, month_num = map(int, month.split('-'))
    days_in_month = calendar.monthrange(year_num, month_num)[1]
    dates = [f"{month}-{str(d).zfill(2)}" for d in range(1, days_in_month + 1)]

    teachers_map = defaultdict(lambda: {'name': '', 'attendance': {}})
    for row in rows:
        tid = row['teacher_id']
        teachers_map[tid]['name'] = row['full_name']
        if row['attendance_date']:
            date_str = str(row['attendance_date'])[:10]
            teachers_map[tid]['attendance'][date_str] = row['status']

    return render_template('teacher_attendance_report.html',
                           teachers_map=dict(teachers_map),
                           dates=dates,
                           month=month)


# ========== STUDENT ATTENDANCE (Admin + Teacher) ==========
@app.route('/attendance/student', methods=['GET', 'POST'])
@login_required
@teacher_required
def student_attendance():
    school_id = session.get('active_school_id', session.get('school_id')) \
        if session['role'] == 'admin' else session.get('school_id')
    today = datetime.date.today()

    conn = get_db()
    c = conn.cursor()

    if session['role'] == 'admin':
        c.execute("SELECT * FROM classes WHERE school_id=%s ORDER BY class_name", (school_id,))
    else:
        c.execute("SELECT id FROM teachers WHERE user_id=%s", (session['user_id'],))
        t = c.fetchone()
        if t:
            c.execute("""
                SELECT c.* FROM classes c
                JOIN teacher_classes tc ON tc.class_id = c.id
                WHERE tc.teacher_id = %s
                ORDER BY c.class_name
            """, (t[0],))
        else:
            conn.close()
            return render_template('student_attendance.html', classes=[], today=today)

    classes_list = fetchall_dict(c)

    selected_class = None
    students = []
    attendance_date = str(today)
    existing = {}

    if request.method == 'POST':
        selected_class = request.form.get('class_id')
        attendance_date = request.form.get('attendance_date', str(today))

        if selected_class:
            c.execute("SELECT * FROM students WHERE class_id=%s ORDER BY full_name", (selected_class,))
            students = fetchall_dict(c)

            c.execute("""
                SELECT student_id, status, remarks
                FROM student_attendance
                WHERE class_id=%s AND attendance_date=%s
            """, (selected_class, attendance_date))
            for row in c.fetchall():
                existing[row[0]] = {'status': row[1], 'remarks': row[2] or ''}

    conn.close()
    return render_template('student_attendance.html',
                           classes=classes_list,
                           selected_class=selected_class,
                           students=students,
                           attendance_date=attendance_date,
                           existing=existing,
                           today=today)


@app.route('/attendance/student/save', methods=['POST'])
@login_required
@teacher_required
def save_student_attendance():
    class_id = request.form.get('class_id')
    attendance_date = request.form.get('attendance_date')
    school_id = session.get('active_school_id', session.get('school_id'))
    marked_by = session['user_id']

    student_ids = request.form.getlist('student_id[]')
    statuses = request.form.getlist('status[]')
    remarks_list = request.form.getlist('remarks[]')

    if not attendance_date or not class_id or not student_ids:
        flash('Class, date aur students ki attendance zaruri hai', 'error')
        return redirect(url_for('student_attendance'))

    conn = get_db()
    c = conn.cursor()

    for i, student_id in enumerate(student_ids):
        status = statuses[i] if i < len(statuses) else 'Absent'
        remarks = remarks_list[i] if i < len(remarks_list) else ''
        # Postgres: INSERT ... ON CONFLICT (upsert) instead of SQL Server IF EXISTS/ELSE
        c.execute("""
            INSERT INTO student_attendance
            (student_id, class_id, school_id, attendance_date,
             status, remarks, marked_by, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (student_id, attendance_date)
            DO UPDATE SET status = EXCLUDED.status,
                          remarks = EXCLUDED.remarks,
                          marked_by = EXCLUDED.marked_by,
                          updated_by = EXCLUDED.marked_by,
                          updated_date = NOW()
        """, (
            student_id, class_id, school_id, attendance_date,
            status, remarks, marked_by, marked_by
        ))

    conn.commit()
    conn.close()
    flash(f'{len(student_ids)} students ki attendance save ho gayi!', 'success')
    return redirect(url_for('student_attendance'))


# ========== FEE COLLECTION ROUTE ==========
@app.route('/fee_collection', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def fee_collection():
    school_id = session.get('active_school_id', session.get('school_id'))
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM classes WHERE school_id=%s ORDER BY class_name", (school_id,))
    classes = fetchall_dict(c)

    if request.method == 'POST':
        class_id = request.form.get('class_id')
        attendance_date = request.form.get('attendance_date', datetime.date.today())

        c.execute("SELECT * FROM students WHERE class_id=%s ORDER BY full_name", (class_id,))
        students = fetchall_dict(c)
        conn.close()

        return render_template('fee_collection.html',
                               classes=classes,
                               selected_class=class_id,
                               students=students,
                               attendance_date=attendance_date,
                               now=datetime.datetime.now())

    conn.close()
    return render_template('fee_collection.html',
                           classes=classes,
                           now=datetime.datetime.now(),
                           students=[])

@app.route('/fee_collection/save', methods=['POST'])
@login_required
@school_admin_only_required
def save_fee():
    student_id = request.form.get('student_id')
    class_id = request.form.get('class_id')
    month = request.form.get('month')
    year = request.form.get('year')
    amount = request.form.get('amount')
    payment_mode = request.form.get('payment_mode')
    transaction_reference = request.form.get('transaction_reference', '')
    remarks = request.form.get('remarks', '')

    school_id = session.get('active_school_id', session.get('school_id'))
    collected_by = session['user_id']

    receipt_number = f"FEE-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"

    conn = get_db()
    c = conn.cursor()
    c.execute("""
        INSERT INTO fee_collections 
        (student_id, school_id, class_id, month, year, amount, 
         payment_mode, transaction_reference, remarks, collected_by,
         receipt_number, created_by)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (student_id, school_id, class_id, month, year, amount,
          payment_mode, transaction_reference, remarks, collected_by,
          receipt_number, session['user_id']))

    fee_id = c.fetchone()[0]
    conn.commit()

    c.execute("""
        SELECT fc.*, s.full_name as student_name, s.student_code,
               c.class_name, u.full_name as collector_name
        FROM fee_collections fc
        JOIN students s ON fc.student_id = s.id
        JOIN classes c ON fc.class_id = c.id
        JOIN users u ON fc.collected_by = u.id
        WHERE fc.id = %s
    """, (fee_id,))
    receipt = fetchone_dict(c)
    conn.close()

    flash(f'Fee collection successful! Receipt: {receipt_number}', 'success')
    return render_template('fee_receipt.html', receipt=receipt)


# Add this debug route temporarily
@app.route('/debug_columns')
def debug_columns():
    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT column_name 
        FROM information_schema.columns 
        WHERE table_name = 'classes'
        ORDER BY ordinal_position
    """)
    columns = c.fetchall()
    conn.close()

    return str([col[0] for col in columns])


@app.route('/teacher/<int:teacher_id>')
def view_teacher(teacher_id):
    teacher = Teacher.query.get_or_404(teacher_id)
    documents = TeacherDocument.query.filter_by(teacher_id=teacher_id).all()
    return render_template('view_teacher.html', teacher=teacher, documents=documents)


@app.route('/get_classes_by_school')
@login_required
def get_classes_by_school():
    school_id = request.args.get('school_id')
    if not school_id:
        return jsonify([])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, class_name, section FROM classes WHERE school_id=%s ORDER BY class_name", (school_id,))
    classes = fetchall_dict(c)
    conn.close()
    return jsonify(classes)


@app.route('/get_teachers_by_school')
@login_required
def get_teachers_by_school():
    school_id = request.args.get('school_id')
    if not school_id:
        return jsonify([])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, full_name, teacher_code FROM teachers WHERE school_id=%s ORDER BY full_name", (school_id,))
    teachers = fetchall_dict(c)
    conn.close()
    return jsonify(teachers)


################################################### notices ########################################
# ========== NOTICE BOARD ROUTES ==========

@app.route('/notices/manage', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def manage_notices():
    school_id = session.get('active_school_id') or session.get('school_id')

    if not school_id:
        flash('School not found!', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()
    c = conn.cursor()

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        body = request.form.get('body', '')
        recipient_role = request.form.get('recipient_role', 'all')

        if not title:
            flash('Title zaroori hai', 'error')
            return redirect(url_for('manage_notices'))

        image_filename = None
        if 'notice_image' in request.files:
            file = request.files['notice_image']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(f"notice_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                os.makedirs('static/uploads/notices', exist_ok=True)
                file.save(os.path.join('static/uploads/notices', filename))
                image_filename = filename

        c.execute("""
            INSERT INTO notices (school_id, title, body, image_path, created_by, created_date)
            VALUES (%s, %s, %s, %s, %s, NOW()) RETURNING id
        """, (school_id, title, body, image_filename, session['user_id']))

        notice_id = c.fetchone()[0]

        c.execute("""
            INSERT INTO notice_recipients (notice_id, role)
            VALUES (%s, %s)
        """, (notice_id, recipient_role))

        conn.commit()
        flash('Notice publish ho gaya!', 'success')
        return redirect(url_for('manage_notices'))

    c.execute("""
        SELECT n.*, u.full_name as creator_name,
               TO_CHAR(n.created_date, 'DD-Mon-YYYY HH12:MI AM') as formatted_date
        FROM notices n
        JOIN users u ON n.created_by = u.id
        WHERE n.school_id = %s
        ORDER BY n.created_date DESC
    """, (school_id,))
    notices = fetchall_dict(c)
    conn.close()

    return render_template('manage_notices.html', notices=notices)


@app.route('/notices')
@login_required
def view_notices():
    school_id = session.get('active_school_id') or session.get('school_id')
    role = session['role']

    if not school_id:
        flash('School not found!', 'error')
        return render_template('view_notices.html', notices=[])

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT n.*, u.full_name as creator_name,
               TO_CHAR(n.created_date, 'DD-Mon-YYYY HH12:MI AM') as formatted_date
        FROM notices n
        JOIN users u ON n.created_by = u.id
        JOIN notice_recipients nr ON n.id = nr.notice_id
        WHERE n.school_id = %s 
          AND n.is_active = TRUE
          AND (nr.role = 'all' OR nr.role = %s)
        ORDER BY n.created_date DESC
    """, (school_id, role))

    notices = fetchall_dict(c)
    conn.close()

    return render_template('view_notices.html', notices=notices)


@app.route('/notices/delete/<int:notice_id>')
@login_required
@school_admin_only_required
def delete_notice(notice_id):
    conn = get_db()
    c = conn.cursor()

    school_id = session.get('active_school_id', session.get('school_id'))
    c.execute("SELECT id FROM notices WHERE id=%s AND school_id=%s", (notice_id, school_id))
    if c.fetchone():
        c.execute("UPDATE notices SET is_active=FALSE WHERE id=%s", (notice_id,))
        conn.commit()
        flash('Notice delete ho gaya!', 'success')
    else:
        flash('Aap is notice ko delete nahi kar sakte', 'error')

    conn.close()
    return redirect(url_for('manage_notices'))


# ========== SCHOOL ADMIN SIGNUP (Public - koi bhi access kar sakta hai) ==========
@app.route('/school_admin_signup', methods=['GET', 'POST'])
def school_admin_signup():
    if 'user_id' in session and session.get('role') == 'admin':
        return redirect(url_for('schools'))

    if request.method == 'POST':
        school_name = request.form.get('name', '').strip()
        school_address = request.form.get('address', '')
        school_phone = request.form.get('phone', '')
        school_email = request.form.get('email', '')
        school_city = request.form.get('city', '')
        country = request.form.get('country', '')
        area = request.form.get('area', '')
        latitude = request.form.get('latitude', '')
        longitude = request.form.get('longitude', '')

        full_name = request.form.get('full_name', '').strip()
        admin_email = request.form.get('admin_email', '') or school_email
        admin_phone = request.form.get('admin_phone', '') or school_phone
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not school_name or not username or not password or not full_name:
            flash('School name, username, password aur full name zaruri hain', 'error')
            return render_template('school_form.html')

        if password != confirm_password:
            flash('Password aur confirm password match nahi ho rahe', 'error')
            return render_template('school_form.html')

        conn = get_db()
        c = conn.cursor()
        try:
            logo = None
            if 'logo' in request.files:
                file = request.files['logo']
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                    file.save(os.path.join(UPLOAD_FOLDER, filename))
                    logo = filename

            c.execute("""
                INSERT INTO schools (name, address, phone, email, city, logo, is_active)
                VALUES (%s, %s, %s, %s, %s, %s, FALSE) RETURNING id
            """, (school_name, school_address, school_phone, school_email, school_city, logo))
            school_id = c.fetchone()[0]

            c.execute("""
                INSERT INTO users
                  (school_id, username, password, role, full_name, email, phone, is_active)
                VALUES (%s, %s, %s, 'school_admin_pending', %s, %s, %s, FALSE) RETURNING id
            """, (
                school_id,
                username,
                hash_password(password),
                full_name,
                admin_email,
                admin_phone
            ))

            conn.commit()
            flash('Registration successful! Super Admin se approval ke baad aap login kar sakenge.', 'success')
            return redirect(url_for('login'))

        except psycopg2.IntegrityError:
            conn.rollback()
            flash('Yeh username pehle se exist karta hai!', 'error')
        except Exception as e:
            conn.rollback()
            flash(f'Error: {str(e)}', 'error')
        finally:
            conn.close()

    return render_template('school_form.html')


# ========== SUPER ADMIN: Pending School Admins Dekhna aur Approve Karna ==========
@app.route('/super/pending_admins')
@login_required
def pending_admins():
    if session.get('role') != 'admin':
        flash('Aap ye page nahi dekh sakte', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT u.id as user_id, u.username, u.full_name, u.email, u.phone, u.created_date,
               s.id as school_id, s.name as school_name, s.address as school_address, 
               s.phone as school_phone, s.email as school_email, s.city as school_city
        FROM users u
        JOIN schools s ON u.school_id = s.id
        WHERE u.role = 'school_admin_pending' AND u.is_active = FALSE
        ORDER BY u.created_date DESC
    """)
    pending_list = fetchall_dict(c)
    conn.close()

    return render_template('pending_admins.html', pending_list=pending_list)


@app.route('/super/approve_admin/<int:user_id>')
@login_required
def approve_admin(user_id):
    if session.get('role') != 'admin':
        flash('Aap approve nahi kar sakte', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()
    c = conn.cursor()

    try:
        c.execute("""
            UPDATE users 
            SET role = 'school_admin', is_active = TRUE, updated_date = NOW()
            WHERE id = %s AND role = 'school_admin_pending'
        """, (user_id,))

        c.execute("""
            UPDATE schools 
            SET is_active = TRUE 
            WHERE id = (SELECT school_id FROM users WHERE id = %s)
        """, (user_id,))

        conn.commit()
        flash('School Admin approved successfully! Ab woh login kar sakta hai.', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('pending_admins'))


@app.route('/super/reject_admin/<int:user_id>')
@login_required
def reject_admin(user_id):
    if session.get('role') != 'admin':
        flash('Aap reject nahi kar sakte', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()
    c = conn.cursor()

    try:
        c.execute("SELECT school_id FROM users WHERE id = %s", (user_id,))
        row = c.fetchone()
        school_id = row[0] if row else None

        c.execute("DELETE FROM users WHERE id = %s AND role = 'school_admin_pending'", (user_id,))

        if school_id:
            c.execute("DELETE FROM schools WHERE id = %s", (school_id,))

        conn.commit()
        flash('School Admin request rejected and deleted!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('pending_admins'))


# ========== PARENT MANAGEMENT (School Admin only) ==========

@app.route('/parents', methods=['GET'])
@login_required
@school_admin_only_required
def parents():
    school_id = session.get('active_school_id') or session.get('school_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT u.id, u.username, u.full_name, u.email, u.phone, u.is_active,
               COUNT(pc.student_id) as child_count
        FROM users u
        LEFT JOIN parent_children pc ON pc.parent_user_id = u.id
        WHERE u.role = 'parent' AND u.school_id = %s
        GROUP BY u.id, u.username, u.full_name, u.email, u.phone, u.is_active
        ORDER BY u.full_name
    """, (school_id,))
    parents_list = fetchall_dict(c)
    conn.close()
    return render_template('parents.html', parents=parents_list)


@app.route('/parents/add', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def add_parent():
    school_id = session.get('active_school_id') or session.get('school_id')
    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT s.id, s.full_name, s.student_code, c.class_name, c.section
        FROM students s
        LEFT JOIN classes c ON s.class_id = c.id
        WHERE s.school_id = %s
        ORDER BY s.full_name
    """, (school_id,))
    students_list = fetchall_dict(c)

    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        child_ids = request.form.getlist('child_ids')

        if not full_name or not username or not password:
            flash('Naam, username aur password zaruri hain', 'error')
            conn.close()
            return render_template('parent_form.html', students=students_list)

        if not child_ids:
            flash('Kam se kam ek child select karein', 'error')
            conn.close()
            return render_template('parent_form.html', students=students_list)

        try:
            c.execute("""
                INSERT INTO users (school_id, username, password, role, full_name, email, phone, is_active)
                VALUES (%s, %s, %s, 'parent', %s, %s, %s, TRUE) RETURNING id
            """, (school_id, username, hash_password(password), full_name, email, phone))
            parent_user_id = c.fetchone()[0]

            for child_id in child_ids:
                c.execute("""
                    INSERT INTO parent_children (parent_user_id, student_id, school_id, created_by)
                    VALUES (%s, %s, %s, %s)
                """, (parent_user_id, child_id, school_id, session['user_id']))

            conn.commit()
            flash(f'Parent login ban gaya! Username: {username}', 'success')
            return redirect(url_for('parents'))

        except Exception as e:
            conn.rollback()
            flash(f'Error: {str(e)}', 'error')
        finally:
            conn.close()

    conn.close()
    return render_template('parent_form.html', students=students_list)


@app.route('/select_child', methods=['GET', 'POST'])
@login_required
def select_child():
    if session['role'] != 'parent':
        return redirect(url_for('dashboard'))

    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT s.id, s.full_name, s.student_code, s.profile_pic,
               cl.class_name, cl.section
        FROM parent_children pc
        JOIN students s ON pc.student_id = s.id
        LEFT JOIN classes cl ON s.class_id = cl.id
        WHERE pc.parent_user_id = %s
        ORDER BY s.full_name
    """, (session['user_id'],))
    children = fetchall_dict(c)
    conn.close()

    if len(children) == 1:
        session['active_child_id'] = children[0]['id']
        return redirect(url_for('parent_dashboard'))

    if request.method == 'POST':
        chosen = request.form.get('child_id')
        if chosen:
            session['active_child_id'] = int(chosen)
            return redirect(url_for('parent_dashboard'))

    return render_template('select_child.html', children=children)


@app.route('/parent/dashboard')
@login_required
def parent_dashboard():
    if session['role'] != 'parent':
        return redirect(url_for('dashboard'))

    child_id = session.get('active_child_id')
    if not child_id:
        return redirect(url_for('select_child'))

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT s.*, cl.class_name, cl.section, sc.name as school_name, sc.logo
        FROM students s
        LEFT JOIN classes cl ON s.class_id = cl.id
        JOIN schools sc ON s.school_id = sc.id
        WHERE s.id = %s
    """, (child_id,))
    child = fetchone_dict(c)

    c.execute("SELECT id FROM parent_children WHERE parent_user_id=%s AND student_id=%s",
              (session['user_id'], child_id))
    if not c.fetchone():
        flash('Aap is child ka data nahi dekh sakte', 'error')
        conn.close()
        return redirect(url_for('select_child'))

    c.execute("""
        SELECT s.id, s.full_name, s.profile_pic
        FROM parent_children pc
        JOIN students s ON pc.student_id = s.id
        WHERE pc.parent_user_id = %s
    """, (session['user_id'],))
    all_children = fetchall_dict(c)

    c.execute("""
        SELECT attendance_date, status
        FROM student_attendance
        WHERE student_id = %s
        ORDER BY attendance_date DESC
        LIMIT 30
    """, (child_id,))
    attendance = fetchall_dict(c)

    c.execute("""
        SELECT month, year, amount, payment_mode, receipt_number, created_date
        FROM fee_collections
        WHERE student_id = %s
        ORDER BY year DESC, month DESC
    """, (child_id,))
    fees = fetchall_dict(c)

    conn.close()
    return render_template('parent_dashboard.html',
                           child=child,
                           all_children=all_children,
                           attendance=attendance,
                           fees=fees)


@app.route('/parent/switch_child/<int:child_id>')
@login_required
def switch_child(child_id):
    if session['role'] != 'parent':
        return redirect(url_for('dashboard'))
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM parent_children WHERE parent_user_id=%s AND student_id=%s",
              (session['user_id'], child_id))
    if c.fetchone():
        session['active_child_id'] = child_id
    conn.close()
    return redirect(url_for('parent_dashboard'))

@app.route('/school_admin/add', methods=['GET', 'POST'])
@login_required
@school_admin_or_super_admin_required
def add_school_admin():
    school_id = session.get('active_school_id') or session.get('school_id')
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT id, name FROM schools WHERE id=%s", (school_id,))
    school = fetchone_dict(c)

    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')

        if session.get('role') == 'admin':
            chosen_school_id = request.form.get('school_id', school_id)
        else:
            chosen_school_id = school_id

        if not full_name or not username or not password:
            flash('Naam, username aur password zaruri hain', 'error')
            conn.close()
            return render_template('add_school_admin.html', school=school)

        try:
            c.execute("""
                INSERT INTO users
                  (school_id, username, password, role, full_name, email, phone, is_active)
                VALUES (%s, %s, %s, 'school_admin', %s, %s, %s, TRUE)
            """, (chosen_school_id, username, hash_password(password), full_name, email, phone))
            conn.commit()
            flash(f'School Admin "{full_name}" ban gaya! Username: {username}', 'success')
            return redirect(url_for('users'))
        except psycopg2.IntegrityError:
            conn.rollback()
            flash('Yeh username pehle se exist karta hai!', 'error')
        except Exception as e:
            conn.rollback()
            flash(f'Error: {str(e)}', 'error')
        finally:
            conn.close()

    conn.close()
    return render_template('add_school_admin.html', school=school)

# CLASSES BULK UPLOAD
@app.route('/bulk/classes', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def bulk_upload_classes():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename or \
           not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Please upload a valid Excel file (.xlsx or .xls)', 'error')
            return redirect(request.url)

        try:
            df = pd.read_excel(file, header=2)
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'error')
            return redirect(request.url)

        df.columns = [str(col).strip().rstrip('*').strip() for col in df.columns]

        conn = get_db()
        c = conn.cursor()
        school_id = session.get('active_school_id') or session.get('school_id')

        c.execute("DELETE FROM staging_classes WHERE school_id=%s AND uploaded_by=%s",
                  (school_id, session['user_id']))

        valid_count = 0
        invalid_count = 0

        for idx, row in df.iterrows():
            row_dict = {str(k).strip().rstrip('*').strip(): v for k, v in row.items()}

            if is_row_empty(row_dict):
                continue

            class_name    = get_col(row_dict, 'Class Name', 'class_name', 'ClassName')
            section       = get_col(row_dict, 'Section', 'section')
            academic_year = get_col(row_dict, 'Academic Year', 'academic_year', 'AcademicYear')

            errors    = validate_classes_row(row_dict, school_id)
            status    = 'VALID' if not errors else 'INVALID'
            error_msg = '; '.join(errors) if errors else None

            c.execute("""
                INSERT INTO staging_classes
                (school_id, class_name, section, academic_year,
                 validation_status, error_message, uploaded_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (school_id, class_name, section, academic_year,
                  status, error_msg, session['user_id']))

            if status == 'VALID':
                valid_count += 1
            else:
                invalid_count += 1

        conn.commit()
        conn.close()
        flash(f'✅ {valid_count} valid | ❌ {invalid_count} invalid rows found', 'info')
        return redirect(url_for('review_classes_upload'))

    return render_template('bulk_upload_updated.html', upload_type='classes')


@app.route('/bulk/teachers', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def bulk_upload_teachers():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename or \
           not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Please upload a valid Excel file (.xlsx or .xls)', 'error')
            return redirect(request.url)

        try:
            df = pd.read_excel(file, header=2)
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'error')
            return redirect(request.url)

        df.columns = [str(col).strip().rstrip('*').strip() for col in df.columns]

        conn = get_db()
        c = conn.cursor()
        school_id = session.get('active_school_id') or session.get('school_id')

        c.execute("DELETE FROM staging_teachers WHERE school_id=%s AND uploaded_by=%s",
                  (school_id, session['user_id']))

        valid_count = 0
        invalid_count = 0

        for idx, row in df.iterrows():
            row_dict = {str(k).strip().rstrip('*').strip(): v for k, v in row.items()}

            if is_row_empty(row_dict):
                continue

            full_name     = get_col(row_dict, 'Full Name', 'full_name', 'FullName')
            username      = get_col(row_dict, 'Username', 'username')
            password      = get_col(row_dict, 'Password', 'password')
            email         = get_col(row_dict, 'Email', 'email')
            phone         = get_col(row_dict, 'Phone', 'phone')
            subject_spec  = get_col(row_dict, 'Subject Specialization', 'subject_specialization')
            qualification = get_col(row_dict, 'Qualification', 'qualification')

            joining_date = get_col(
                row_dict,
                'Joining Date',
                'joining_date',
                'JoiningDate'
            )

            if pd.notna(joining_date) and joining_date:
                try:
                    joining_date = pd.to_datetime(joining_date).date()
                except Exception:
                    joining_date = None

            errors    = validate_teacher_row(row_dict, conn)
            status    = 'VALID' if not errors else 'INVALID'
            error_msg = '; '.join(errors) if errors else None

            hashed_pwd = hash_password(password) if password else ''

            c.execute("""
                INSERT INTO staging_teachers
                (school_id, full_name, email, phone, subject_specialization,
                 qualification, joining_date, username, password,
                 validation_status, error_message, uploaded_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                school_id,
                full_name,
                email,
                phone,
                subject_spec,
                qualification,
                joining_date,
                username,
                hashed_pwd,
                status,
                error_msg,
                session['user_id']
            ))

            if status == 'VALID':
                valid_count += 1
            else:
                invalid_count += 1

        conn.commit()
        conn.close()
        flash(f'✅ {valid_count} valid | ❌ {invalid_count} invalid rows found', 'info')
        return redirect(url_for('review_teachers_upload'))

    return render_template('bulk_upload_updated.html', upload_type='teachers')

@app.route('/bulk/students', methods=['GET', 'POST'])
@login_required
@school_admin_only_required
def bulk_upload_students():

    from datetime import datetime

    def clean_date(dob):
        try:
            if not dob or str(dob).strip() == "":
                return None

            return datetime.strptime(str(dob), "%Y-%m-%d").date()
        except:
            try:
                return datetime.strptime(str(dob), "%d/%m/%Y").date()
            except:
                return None

    if request.method == 'POST':

        file = request.files.get('file')

        if not file or not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Please upload a valid Excel file (.xlsx or .xls)', 'error')
            return redirect(request.url)

        try:
            df = pd.read_excel(file, header=2)
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'error')
            return redirect(request.url)

        df.columns = [str(col).strip().rstrip('*').strip() for col in df.columns]

        conn = get_db()
        c = conn.cursor()

        school_id = session.get('active_school_id') or session.get('school_id')

        c.execute("""
            DELETE FROM staging_students
            WHERE school_id=%s AND uploaded_by=%s
        """, (school_id, session['user_id']))

        valid_count = 0
        invalid_count = 0

        for idx, row in df.iterrows():

            row_dict = {str(k).strip().rstrip('*').strip(): v for k, v in row.items()}

            if is_row_empty(row_dict):
                continue

            class_id_str = get_col(row_dict, 'Class ID', 'class_id', 'ClassID')
            full_name    = get_col(row_dict, 'Full Name', 'full_name', 'FullName')
            father_name  = get_col(row_dict, 'Father Name', 'father_name', 'FatherName')
            email        = get_col(row_dict, 'Email', 'email')
            phone        = get_col(row_dict, 'Phone', 'phone')
            gender       = get_col(row_dict, 'Gender', 'gender')
            username     = get_col(row_dict, 'Username', 'username')
            password     = get_col(row_dict, 'Password', 'password')

            dob = get_col(row_dict, 'Date of Birth', 'date_of_birth', 'DOB')
            dob = clean_date(dob)

            class_id_db = None
            if class_id_str:
                try:
                    class_id_db = int(float(class_id_str))
                except:
                    class_id_db = None

            errors = validate_student_row(row_dict, conn)

            status = 'VALID' if not errors else 'INVALID'
            error_msg = '; '.join(errors) if errors else None

            hashed_pwd = hash_password(password) if password else ''

            c.execute("""
                INSERT INTO staging_students
                (school_id, class_id, full_name, father_name, email, phone,
                 date_of_birth, gender, username, password,
                 validation_status, error_message, uploaded_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                school_id,
                class_id_db,
                full_name,
                father_name,
                email,
                phone,
                dob,
                gender,
                username,
                hashed_pwd,
                status,
                error_msg,
                session['user_id']
            ))

            if status == 'VALID':
                valid_count += 1
            else:
                invalid_count += 1

        conn.commit()
        conn.close()

        flash(f'✅ {valid_count} valid | ❌ {invalid_count} invalid rows found', 'info')
        return redirect(url_for('review_students_upload'))

    conn = get_db()
    c = conn.cursor()

    school_id = session.get('active_school_id') or session.get('school_id')

    c.execute("""
        SELECT id, class_name, section
        FROM classes
        WHERE school_id=%s
        ORDER BY class_name
    """, (school_id,))

    classes_list = fetchall_dict(c)
    conn.close()

    return render_template(
        'bulk_upload_updated.html',
        upload_type='students',
        classes=classes_list
    )


@app.route('/bulk/<upload_type>/download_report')
@login_required
@school_admin_only_required
def download_validation_report(upload_type):
    """Download staging table as Excel validation report."""
    from flask import send_file
    import io
    conn = get_db()
    c = conn.cursor()
    school_id = session.get('active_school_id') or session.get('school_id')

    table_map = {
        'classes': 'staging_classes',
        'teachers': 'staging_teachers',
        'students': 'staging_students',
    }
    table = table_map.get(upload_type)
    if not table:
        flash('Invalid upload type', 'error')
        return redirect(url_for('dashboard'))

    c.execute(f"SELECT * FROM {table} WHERE school_id=%s AND uploaded_by=%s ORDER BY id",
              (school_id, session['user_id']))
    records = fetchall_dict(c)
    conn.close()

    excel_bytes = generate_excel_report(records, upload_type)
    return send_file(
        io.BytesIO(excel_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{upload_type}_validation_report.xlsx'
    )

# ========== REVIEW AND CONFIRM ROUTES ==========

@app.route('/bulk/classes/review')
@login_required
@school_admin_only_required
def review_classes_upload():
    conn = get_db()
    c = conn.cursor()
    school_id = session.get('active_school_id') or session.get('school_id')

    c.execute("""
        SELECT * FROM staging_classes 
        WHERE school_id=%s AND uploaded_by=%s
        ORDER BY id
    """, (school_id, session['user_id']))
    records = fetchall_dict(c)
    conn.close()

    return render_template('review_upload.html', records=records, upload_type='classes')

@app.route('/bulk/classes/confirm')
@login_required
@school_admin_only_required
def confirm_classes_upload():
    conn = get_db()
    c = conn.cursor()
    school_id = session.get('active_school_id') or session.get('school_id')

    c.execute("""
        INSERT INTO classes (school_id, class_name, section, academic_year)
        SELECT school_id, class_name, section, academic_year
        FROM staging_classes
        WHERE school_id=%s AND validation_status='VALID' AND uploaded_by=%s
    """, (school_id, session['user_id']))

    c.execute("DELETE FROM staging_classes WHERE school_id=%s AND uploaded_by=%s",
              (school_id, session['user_id']))

    conn.commit()
    conn.close()
    flash('Classes uploaded successfully!', 'success')
    return redirect(url_for('classes'))

@app.route('/bulk/teachers/review')
@login_required
@school_admin_only_required
def review_teachers_upload():
    conn = get_db()
    c = conn.cursor()
    school_id = session.get('active_school_id') or session.get('school_id')

    c.execute("""
        SELECT * FROM staging_teachers 
        WHERE school_id=%s AND uploaded_by=%s
        ORDER BY id
    """, (school_id, session['user_id']))
    records = fetchall_dict(c)
    conn.close()

    return render_template('review_upload.html', records=records, upload_type='teachers')

@app.route('/bulk/teachers/confirm')
@login_required
@school_admin_only_required
def confirm_teachers_upload():
    conn = get_db()
    c = conn.cursor()
    school_id = session.get('active_school_id') or session.get('school_id')

    c.execute("""
        SELECT id, username, full_name, email, phone, password,
               subject_specialization, qualification, joining_date
        FROM staging_teachers
        WHERE school_id=%s AND validation_status='VALID' AND uploaded_by=%s
    """, (school_id, session['user_id']))
    staging_records = c.fetchall()

    for record in staging_records:
        staging_id, username, full_name, email, phone, pwd, subj, qual, joining = record

        c.execute("""
            INSERT INTO users (school_id, username, password, role, full_name, email, phone, is_active)
            VALUES (%s,%s,%s,'teacher',%s,%s,%s,TRUE) RETURNING id
        """, (school_id, username, pwd, full_name, email, phone))
        user_id = c.fetchone()[0]

        teacher_code = f"TCH-{staging_id}"
        c.execute("""
            INSERT INTO teachers (user_id, school_id, teacher_code, full_name, email, phone,
                                  subject_specialization, qualification, joining_date)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (user_id, school_id, teacher_code, full_name, email, phone, subj, qual, joining))

    c.execute("DELETE FROM staging_teachers WHERE school_id=%s AND uploaded_by=%s",
              (school_id, session['user_id']))
    conn.commit()
    conn.close()
    flash('Teachers uploaded successfully!', 'success')
    return redirect(url_for('teachers'))

@app.route('/bulk/students/review')
@login_required
@school_admin_only_required
def review_students_upload():
    conn = get_db()
    c = conn.cursor()
    school_id = session.get('active_school_id') or session.get('school_id')

    c.execute("""
        SELECT * FROM staging_students 
        WHERE school_id=%s AND uploaded_by=%s
        ORDER BY id
    """, (school_id, session['user_id']))
    records = fetchall_dict(c)
    conn.close()

    return render_template('review_upload.html', records=records, upload_type='students')


@app.route('/bulk/students/confirm')
@login_required
@school_admin_only_required
def confirm_students_upload():
    conn = get_db()
    c = conn.cursor()
    school_id = session.get('active_school_id') or session.get('school_id')

    c.execute("""
        SELECT id, username, full_name, email, phone, password,
               class_id, father_name, date_of_birth, gender
        FROM staging_students
        WHERE school_id=%s AND validation_status='VALID' AND uploaded_by=%s
    """, (school_id, session['user_id']))
    staging_records = c.fetchall()

    for record in staging_records:
        staging_id, username, full_name, email, phone, pwd, class_id, father, dob, gender = record

        c.execute("""
            INSERT INTO users (school_id, username, password, role, full_name, email, phone, is_active)
            VALUES (%s,%s,%s,'student',%s,%s,%s,TRUE) RETURNING id
        """, (school_id, username, pwd, full_name, email, phone))
        user_id = c.fetchone()[0]

        student_code = f"STD-{staging_id}"
        c.execute("""
            INSERT INTO students (user_id, school_id, student_code, full_name, father_name,
                                  email, phone, date_of_birth, gender, class_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (user_id, school_id, student_code, full_name, father, email, phone, dob, gender, class_id))

    c.execute("DELETE FROM staging_students WHERE school_id=%s AND uploaded_by=%s",
              (school_id, session['user_id']))
    conn.commit()
    conn.close()
    flash('Students uploaded successfully!', 'success')
    return redirect(url_for('students'))


# ════════════════════════════════════════════════════════════════
#  Excel template generator helpers (unaffected by DB engine)
# ════════════════════════════════════════════════════════════════

import io
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def make_border():
    thin = Side(style='thin', color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_cell(cell, value="", bold=False, fg="000000", bg="FFFFFF",
               center=False, wrap=False, size=10):
    cell.value = value
    cell.font = Font(bold=bold, color=fg, name="Arial", size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.border = make_border()
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=wrap
    )


def add_title_row(ws, text, num_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    style_cell(ws.cell(1, 1), text, bold=True, fg="FFFFFF", bg="1F4E79",
               center=True, size=14)
    ws.row_dimensions[1].height = 32


def add_sequence_row(ws, text, num_cols):
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    style_cell(ws.cell(2, 1), text, bold=True, fg="FF0000", bg="FFFF00",
               wrap=True, size=10)
    ws.row_dimensions[2].height = 30


def add_legend_row(ws, num_cols):
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
    style_cell(ws.cell(3, 1),
               "RED header = MANDATORY (required)   |   GREEN header = Optional field   |   Do NOT change column order",
               fg="000000", bg="F2F2F2", size=9, wrap=True)
    ws.row_dimensions[3].height = 18


def add_headers(ws, columns, row=4):
    RED_BG   = "FF0000"
    GREEN_BG = "00B050"
    for col_idx, (label, mandatory) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx)
        style_cell(cell, label,
                   bold=True, fg="FFFFFF",
                   bg=RED_BG if mandatory else GREEN_BG,
                   center=True)
    ws.row_dimensions[row].height = 22


def add_sample_rows(ws, rows, start_row=5):
    for r_idx, row_data in enumerate(rows):
        bg = "F2F2F2" if r_idx % 2 == 0 else "FFFFFF"
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=10, color="000000")
            cell.fill = PatternFill("solid", start_color=bg)
            cell.border = make_border()
            cell.alignment = Alignment(horizontal="left", vertical="center")


def add_instructions_sheet(wb, fields):
    ns = wb.create_sheet("Instructions")
    headers = ["Column Name", "Description", "Mandatory", "Example Value"]
    for ci, h in enumerate(headers, 1):
        style_cell(ns.cell(1, ci), h, bold=True, fg="FFFFFF",
                   bg="1F4E79", center=True)

    for ri, (name, desc, mandatory, example) in enumerate(fields, 2):
        ns.cell(ri, 1).value = name
        ns.cell(ri, 2).value = desc
        ns.cell(ri, 3).value = "YES" if mandatory else "No"
        ns.cell(ri, 4).value = example

        mand_color = "FFE0E0" if mandatory else "E0FFE0"
        for ci in range(1, 5):
            cell = ns.cell(ri, ci)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=mand_color)
            cell.border = make_border()
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ns.row_dimensions[ri].height = 28

    ns.column_dimensions['A'].width = 25
    ns.column_dimensions['B'].width = 45
    ns.column_dimensions['C'].width = 12
    ns.column_dimensions['D'].width = 30

@app.route('/template/classes')
@login_required
def download_classes_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    school_id = session.get('active_school_id') or session.get('school_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name FROM schools WHERE id=%s", (school_id,))
    row = c.fetchone()
    conn.close()
    school_name = row[0] if row else 'Your School'

    wb = Workbook()
    ws = wb.active
    ws.title = "Classes"

    def bd():
        t = Side(style='thin', color='000000')
        return Border(left=t, right=t, top=t, bottom=t)

    def sc(cell, val='', bold=False, fg='000000', bg='FFFFFF', center=False, sz=10, wrap=False):
        cell.value = val
        cell.font = Font(bold=bold, color=fg, name='Arial', size=sz)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.border = bd()
        cell.alignment = Alignment(horizontal='center' if center else 'left',
                                   vertical='center', wrap_text=wrap)

    ws.merge_cells('A1:D1')
    sc(ws['A1'], f'CLASSES TEMPLATE — {school_name}  (School ID: {school_id})',
       bold=True, fg='FFFFFF', bg='1F4E79', center=True, sz=12)
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:D2')
    sc(ws['A2'], '✅  School ID already filled (blue column). Sirf Class Name, Section bharein.',
       bold=True, fg='7B3F00', bg='FFF2CC', center=True)
    ws.row_dimensions[2].height = 20

    hdrs = [('School ID', True), ('Class Name', True), ('Section', True), ('Academic Year', False)]
    for ci, (lbl, req) in enumerate(hdrs, 1):
        sc(ws.cell(3, ci), lbl, bold=True, fg='FFFFFF',
           bg='C00000' if req else '375623', center=True)
    ws.row_dimensions[3].height = 20

    for r in range(4, 24):
        alt = 'F2F2F2' if r % 2 == 0 else 'FFFFFF'
        a = ws.cell(r, 1, value=school_id)
        a.font = Font(name='Arial', size=10, color='1F4E79', bold=True)
        a.fill = PatternFill('solid', start_color='D9E1F2')
        a.border = bd()
        a.alignment = Alignment(horizontal='center', vertical='center')
        for ci in range(2, 5):
            cell = ws.cell(r, ci)
            cell.font = Font(name='Arial', size=10)
            cell.fill = PatternFill('solid', start_color=alt)
            cell.border = bd()

    for i, w in enumerate([14, 22, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'classes_template_{school_id}.xlsx')


@app.route('/template/teachers')
@login_required
def download_teachers_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    school_id = session.get('active_school_id') or session.get('school_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name FROM schools WHERE id=%s", (school_id,))
    row = c.fetchone()
    conn.close()
    school_name = row[0] if row else 'Your School'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Teachers'

    def bd():
        t = Side(style='thin', color='000000')
        return Border(left=t, right=t, top=t, bottom=t)

    def sc(cell, val='', bold=False, fg='000000', bg='FFFFFF', center=False, sz=10, wrap=False):
        cell.value = val
        cell.font = Font(bold=bold, color=fg, name='Arial', size=sz)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.border = bd()
        cell.alignment = Alignment(horizontal='center' if center else 'left',
                                   vertical='center', wrap_text=wrap)

    num_cols = 9
    last_col = get_column_letter(num_cols)

    ws.merge_cells(f'A1:{last_col}1')
    sc(ws['A1'], f'TEACHERS TEMPLATE — {school_name}  (School ID: {school_id})',
       bold=True, fg='FFFFFF', bg='1F4E79', center=True, sz=12)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{last_col}2')
    sc(ws['A2'], '✅  School ID already filled (blue column). Remaining fields bharein.',
       bold=True, fg='7B3F00', bg='FFF2CC', center=True)
    ws.row_dimensions[2].height = 20

    hdrs = [
        ('School ID', True), ('Full Name', True), ('Username', True), ('Password', True),
        ('Email', False), ('Phone', False), ('Subject Specialization', False),
        ('Qualification', False), ('Joining Date', False),
    ]
    for ci, (lbl, req) in enumerate(hdrs, 1):
        sc(ws.cell(3, ci), lbl, bold=True, fg='FFFFFF',
           bg='C00000' if req else '375623', center=True)
    ws.row_dimensions[3].height = 20

    for r in range(4, 24):
        alt = 'F2F2F2' if r % 2 == 0 else 'FFFFFF'
        a = ws.cell(r, 1, value=school_id)
        a.font = Font(name='Arial', size=10, color='1F4E79', bold=True)
        a.fill = PatternFill('solid', start_color='D9E1F2')
        a.border = bd()
        a.alignment = Alignment(horizontal='center', vertical='center')
        for ci in range(2, num_cols + 1):
            cell = ws.cell(r, ci)
            cell.font = Font(name='Arial', size=10)
            cell.fill = PatternFill('solid', start_color=alt)
            cell.border = bd()

    col_widths = [14, 22, 18, 14, 24, 16, 26, 18, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'teachers_template_{school_id}.xlsx')

@app.route('/template/students')
@login_required
def download_students_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    school_id = session.get('active_school_id') or session.get('school_id')

    selected_class_id = request.args.get('class_id', '').strip()
    selected_class_label = None

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name FROM schools WHERE id=%s", (school_id,))
    row = c.fetchone()

    if selected_class_id:
        c.execute("SELECT id, class_name, section FROM classes WHERE id=%s AND school_id=%s",
                  (selected_class_id, school_id))
        crow = c.fetchone()
        if crow:
            selected_class_id = crow[0]
            selected_class_label = f"{crow[1]} - {crow[2]}"
        else:
            selected_class_id = None

    c.execute("SELECT id, class_name, section FROM classes WHERE school_id=%s ORDER BY class_name, section",
              (school_id,))
    classes = c.fetchall()
    conn.close()
    school_name = row[0] if row else 'Your School'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Students'

    def bd():
        t = Side(style='thin', color='000000')
        return Border(left=t, right=t, top=t, bottom=t)

    def sc(cell, val='', bold=False, fg='000000', bg='FFFFFF', center=False, sz=10, wrap=False):
        cell.value = val
        cell.font = Font(bold=bold, color=fg, name='Arial', size=sz)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.border = bd()
        cell.alignment = Alignment(horizontal='center' if center else 'left',
                                   vertical='center', wrap_text=wrap)

    num_cols = 12
    last_col = get_column_letter(num_cols)

    ws.merge_cells(f'A1:{last_col}1')
    sc(ws['A1'], f'STUDENTS TEMPLATE — {school_name}  (School ID: {school_id})',
       bold=True, fg='FFFFFF', bg='1F4E79', center=True, sz=12)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{last_col}2')
    if selected_class_id:
        note = f'✅  School ID aur Class ID ({selected_class_label}) already filled. Baqi fields bharein.'
    else:
        note = '✅  School ID already filled. Class ID dekhnay ke liye "Class Reference" sheet dekhein.'
    sc(ws['A2'], note, bold=True, fg='7B3F00', bg='FFF2CC', center=True)
    ws.row_dimensions[2].height = 20

    hdrs = [
        ('School ID', True), ('Class ID', True), ('Full Name', True),
        ('Username', True), ('Password', True), ('Father Name', False),
        ('Email', False), ('Phone', False), ('Date of Birth', False),
        ('Gender', False), ('Address', False), ('City', False),
    ]
    for ci, (lbl, req) in enumerate(hdrs, 1):
        sc(ws.cell(3, ci), lbl, bold=True, fg='FFFFFF',
           bg='C00000' if req else '375623', center=True)
    ws.row_dimensions[3].height = 20

    for r in range(4, 24):
        alt = 'F2F2F2' if r % 2 == 0 else 'FFFFFF'

        a = ws.cell(r, 1, value=school_id)
        a.font = Font(name='Arial', size=10, color='1F4E79', bold=True)
        a.fill = PatternFill('solid', start_color='D9E1F2')
        a.border = bd()
        a.alignment = Alignment(horizontal='center', vertical='center')

        start_col = 2
        if selected_class_id:
            b = ws.cell(r, 2, value=selected_class_id)
            b.font = Font(name='Arial', size=10, color='1F4E79', bold=True)
            b.fill = PatternFill('solid', start_color='D9E1F2')
            b.border = bd()
            b.alignment = Alignment(horizontal='center', vertical='center')
            start_col = 3

        for ci in range(start_col, num_cols + 1):
            cell = ws.cell(r, ci)
            cell.font = Font(name='Arial', size=10)
            cell.fill = PatternFill('solid', start_color=alt)
            cell.border = bd()

    col_widths = [13, 12, 22, 18, 14, 20, 24, 16, 16, 10, 24, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if not selected_class_id:
        ref = wb.create_sheet('Class Reference')
        ref.merge_cells('A1:C1')
        if classes:
            ref.cell(1, 1).value = f'{school_name} — Yahan se Class ID copy karein'
            ref.cell(1, 1).font = Font(bold=True, color='FFFFFF', name='Arial')
            ref.cell(1, 1).fill = PatternFill('solid', start_color='1F4E79')
            ref.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
            ref.row_dimensions[1].height = 25

            for ci, h in enumerate(['Class ID (copy this)', 'Class Name', 'Section'], 1):
                cell = ref.cell(2, ci, value=h)
                cell.font = Font(bold=True, color='FFFFFF', name='Arial')
                cell.fill = PatternFill('solid', start_color='375623')
                cell.border = bd()
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for ri, (cid, cname, sec) in enumerate(classes, 3):
                alt2 = 'F2F2F2' if ri % 2 == 0 else 'FFFFFF'
                for ci, val in enumerate([cid, cname, sec], 1):
                    cell = ref.cell(ri, ci, value=val)
                    cell.font = Font(name='Arial', size=10, bold=(ci == 1))
                    cell.fill = PatternFill('solid', start_color=alt2)
                    cell.border = bd()
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            ref.cell(1, 1).value = '⚠️ Koi class nahi mili! Pehle Classes upload karein, phir yeh template dobara download karein.'
            ref.cell(1, 1).font = Font(bold=True, color='C00000', name='Arial')
            ref.cell(1, 1).fill = PatternFill('solid', start_color='FFFF00')
            ref.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ref.row_dimensions[1].height = 40

        for i, w in enumerate([22, 22, 15], 1):
            ref.column_dimensions[get_column_letter(i)].width = w

    fname_suffix = f"_class{selected_class_id}" if selected_class_id else ""

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'students_template_{school_id}{fname_suffix}.xlsx')

if __name__ == "__main__":
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
