"""
reports.py - Excel Export Routes
School Management System

4 Reports:
  1. Student Profile Report (single student OR all students by school/class)
  2. Teacher Report (all teachers by school)
  3. Fee Collection Report (by school + class + month)
  4. Teacher Salary Report (by school + month)
"""

from flask import Blueprint, request, send_file, session, redirect, url_for, flash, render_template
from functools import wraps
from io import BytesIO
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db, fetchall_dict, fetchone_dict

reports_bp = Blueprint('reports', __name__)

# ── helpers ──────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Pehle login karein', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Sirf Admin access kar sakta hai', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

  
def teacher_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') not in ['admin', 'teacher']:
            flash('Aapko is page ka access nahi hai', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

# Style helpers
HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
SUBHEAD_FILL  = PatternFill("solid", start_color="2E75B6")
ALT_FILL      = PatternFill("solid", start_color="DEEAF1")
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT    = Font(name="Arial", bold=True, size=14)
NORMAL_FONT   = Font(name="Arial", size=10)
BOLD_FONT     = Font(name="Arial", bold=True, size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="B8CCE4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, fill=None):
    cell.font   = HEADER_FONT
    cell.fill   = fill or HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

def style_cell(cell, bold=False, alt=False):
    cell.font      = BOLD_FONT if bold else NORMAL_FONT
    cell.fill      = ALT_FILL if alt else WHITE_FILL
    cell.alignment = LEFT
    cell.border    = BORDER

def auto_col_width(ws, extra=4):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + extra, 50)

def excel_response(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

# ── REPORT PAGES (GET: form, POST: download) ─────────────────────────────────

@reports_bp.route('/reports')
@login_required
def reports_index():
    conn = get_db()
    c    = conn.cursor()
    school_id = session.get('active_school_id', session.get('school_id'))

    c.execute("SELECT id, name FROM schools ORDER BY name")
    schools = fetchall_dict(c)

    c.execute("SELECT id, class_name, section FROM classes WHERE school_id=%s ORDER BY class_name", (school_id,))
    classes = fetchall_dict(c)

    c.execute("SELECT id, full_name FROM teachers WHERE school_id=%s ORDER BY full_name", (school_id,))
    teachers = fetchall_dict(c)

    conn.close()
    return render_template('reports.html',
                           schools=schools,
                           classes=classes,
                           teachers=teachers,
                           school_id=school_id,
                           now=datetime.datetime.now())


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT 1 – STUDENT PROFILE REPORT
# ═══════════════════════════════════════════════════════════════════════════════

@reports_bp.route('/reports/students/export')
@login_required
@teacher_required
def export_students():
    school_id = request.args.get('school_id') or session.get('active_school_id') or session.get('school_id')
    class_id  = request.args.get('class_id')
    student_id_single = request.args.get('student_id')

    conn = get_db()
    c    = conn.cursor()

    # School info
    c.execute("SELECT name FROM schools WHERE id=%s", (school_id,))
    school_row = c.fetchone()
    school_name = school_row[0] if school_row else "School"

    if student_id_single:
        # Single student
        c.execute("""
            SELECT st.*, c.class_name, c.section, s.name AS school_name
            FROM students st
            LEFT JOIN classes c ON st.class_id = c.id
            JOIN schools s ON st.school_id = s.id
            WHERE st.id = %s
        """, (student_id_single,))
        students = fetchall_dict(c)
    elif class_id:
        c.execute("""
            SELECT st.*, c.class_name, c.section, s.name AS school_name
            FROM students st
            LEFT JOIN classes c ON st.class_id = c.id
            JOIN schools s ON st.school_id = s.id
            WHERE st.school_id = %s AND st.class_id = %s
            ORDER BY st.full_name
        """, (school_id, class_id))
        students = fetchall_dict(c)
    else:
        c.execute("""
            SELECT st.*, c.class_name, c.section, s.name AS school_name
            FROM students st
            LEFT JOIN classes c ON st.class_id = c.id
            JOIN schools s ON st.school_id = s.id
            WHERE st.school_id = %s
            ORDER BY c.class_name, st.full_name
        """, (school_id,))
        students = fetchall_dict(c)

    # Fetch parents & siblings for each student
    all_parents  = {}
    all_siblings = {}
    for s in students:
        sid = s['id']
        c.execute("SELECT * FROM student_parents WHERE student_id=%s", (sid,))
        all_parents[sid] = fetchall_dict(c)
        c.execute("""
            SELECT ss.*, st2.full_name as linked_name
            FROM student_siblings ss
            LEFT JOIN students st2 ON ss.sibling_student_id = st2.id
            WHERE ss.student_id = %s
        """, (sid,))
        all_siblings[sid] = fetchall_dict(c)

    conn.close()

    wb = openpyxl.Workbook()

    # ── Summary Sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Students Summary"
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A1:N1")
    ws["A1"] = f"{school_name} – Student Report"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:N2")
    ws["A2"] = f"Generated: {datetime.datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws["A2"].font = NORMAL_FONT
    ws["A2"].alignment = CENTER

    headers = ["#", "Student Code", "Full Name", "Father Name", "Gender",
               "Date of Birth", "Class", "Section", "Phone", "Email",
               "City", "Joining Date", "Medical Details", "Address"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        style_header(cell)

    for i, s in enumerate(students, 1):
        row = i + 4
        alt = (i % 2 == 0)
        vals = [
            i,
            s.get('student_code', ''),
            s.get('full_name', ''),
            s.get('father_name', ''),
            s.get('gender', ''),
            str(s.get('date_of_birth', '') or ''),
            s.get('class_name', ''),
            s.get('section', ''),
            s.get('phone', ''),
            s.get('email', ''),
            s.get('city', ''),
            str(s.get('joining_date', '') or ''),
            s.get('medical_details', ''),
            f"{s.get('address_line1','')} {s.get('address_line2','')}".strip(),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            style_cell(cell, alt=alt)

    auto_col_width(ws)

    # ── Individual Detail Sheets (one per student if single or ≤30) ──────────
    if len(students) <= 30:
        for s in students:
            title = f"{s.get('full_name','Student')[:25]}"
            ws2 = wb.create_sheet(title=title)

            # Title
            ws2.merge_cells("A1:D1")
            ws2["A1"] = f"Student Profile – {s.get('full_name','')}"
            ws2["A1"].font = TITLE_FONT
            ws2["A1"].alignment = CENTER
            ws2.row_dimensions[1].height = 25

            def add_row(label, value, row_num, alt=False):
                c1 = ws2.cell(row=row_num, column=1, value=label)
                c2 = ws2.cell(row=row_num, column=2, value=str(value or ''))
                style_cell(c1, bold=True, alt=alt)
                style_cell(c2, alt=alt)

            fields = [
                ("Student Code",    s.get('student_code')),
                ("Full Name",       s.get('full_name')),
                ("Father Name",     s.get('father_name')),
                ("Gender",          s.get('gender')),
                ("Date of Birth",   s.get('date_of_birth')),
                ("Phone",           s.get('phone')),
                ("Email",           s.get('email')),
                ("Class",           f"{s.get('class_name','')} {s.get('section','')}".strip()),
                ("Joining Date",    s.get('joining_date')),
                ("Address",         f"{s.get('address_line1','')} {s.get('address_line2','')}".strip()),
                ("City",            s.get('city')),
                ("State",           s.get('state')),
                ("Postal Code",     s.get('postal_code')),
                ("Medical Details", s.get('medical_details')),
            ]
            for idx, (lbl, val) in enumerate(fields):
                add_row(lbl, val, idx + 2, alt=(idx % 2 == 0))

            next_row = len(fields) + 3

            # Parents
            parents = all_parents.get(s['id'], [])
            if parents:
                ws2.cell(row=next_row, column=1, value="Parents / Guardians").font = SUBHEAD_FONT
                ws2.cell(row=next_row, column=1).fill = SUBHEAD_FILL
                ws2.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
                next_row += 1
                p_headers = ["Name", "Relation", "Occupation", "Phone", "Email"]
                for ci, ph in enumerate(p_headers, 1):
                    cell = ws2.cell(row=next_row, column=ci, value=ph)
                    style_header(cell)
                next_row += 1
                for pi, p in enumerate(parents):
                    alt = (pi % 2 == 0)
                    for ci, v in enumerate([p.get('parent_name'), p.get('relation'), p.get('occupation'), p.get('phone'), p.get('email')], 1):
                        style_cell(ws2.cell(row=next_row, column=ci, value=str(v or '')), alt=alt)
                    next_row += 1

            # Siblings
            siblings = all_siblings.get(s['id'], [])
            if siblings:
                next_row += 1
                ws2.cell(row=next_row, column=1, value="Siblings").font = SUBHEAD_FONT
                ws2.cell(row=next_row, column=1).fill = SUBHEAD_FILL
                ws2.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
                next_row += 1
                sib_headers = ["Name", "Age", "Class", "School"]
                for ci, sh in enumerate(sib_headers, 1):
                    cell = ws2.cell(row=next_row, column=ci, value=sh)
                    style_header(cell)
                next_row += 1
                for si2, sib in enumerate(siblings):
                    alt = (si2 % 2 == 0)
                    for ci, v in enumerate([sib.get('sibling_name'), sib.get('age'), sib.get('class'), sib.get('school_name')], 1):
                        style_cell(ws2.cell(row=next_row, column=ci, value=str(v or '')), alt=alt)
                    next_row += 1

            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 30
            ws2.column_dimensions['C'].width = 20
            ws2.column_dimensions['D'].width = 25

    fname = f"Students_Report_{school_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return excel_response(wb, fname)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT 2 – TEACHER REPORT
# ═══════════════════════════════════════════════════════════════════════════════

@reports_bp.route('/reports/teachers/export')
@login_required
@admin_required
def export_teachers():
    school_id = request.args.get('school_id') or session.get('active_school_id') or session.get('school_id')

    conn = get_db()
    c    = conn.cursor()

    c.execute("SELECT name FROM schools WHERE id=%s", (school_id,))
    school_row  = c.fetchone()
    school_name = school_row[0] if school_row else "School"

    c.execute("""
        SELECT t.*,
               (SELECT STRING_AGG(c.class_name || ' ' || c.section, ', ')
                FROM teacher_classes tc
                JOIN classes c ON tc.class_id = c.id
                WHERE tc.teacher_id = t.id) AS assigned_classes
        FROM teachers t
        WHERE t.school_id = %s
        ORDER BY t.full_name
    """, (school_id,))
    teachers = fetchall_dict(c)

    # Fetch documents per teacher
    all_docs = {}
    for t in teachers:
        c.execute("SELECT document_type, document_name FROM teacher_documents WHERE teacher_id=%s", (t['id'],))
        all_docs[t['id']] = fetchall_dict(c)

    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teachers"

    ws.merge_cells("A1:L1")
    ws["A1"] = f"{school_name} – Teacher Report"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Generated: {datetime.datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws["A2"].font = NORMAL_FONT
    ws["A2"].alignment = CENTER

    headers = ["#", "Teacher Code", "Full Name", "CNIC", "Phone", "Email",
               "Qualification", "Specialization", "Joining Date",
               "Salary", "Address", "Assigned Classes"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=4, column=col, value=h))

    for i, t in enumerate(teachers, 1):
        row = i + 4
        alt = (i % 2 == 0)
        vals = [
            i,
            t.get('teacher_code', ''),
            t.get('full_name', ''),
            t.get('cnic', ''),
            t.get('phone', ''),
            t.get('email', ''),
            t.get('qualification', ''),
            t.get('subject_specialization', ''),
            str(t.get('joining_date', '') or ''),
            t.get('salary', ''),
            t.get('address', ''),
            t.get('assigned_classes', ''),
        ]
        for col, v in enumerate(vals, 1):
            style_cell(ws.cell(row=row, column=col, value=v), alt=alt)

    auto_col_width(ws)

    # Documents sheet
    ws2 = wb.create_sheet("Teacher Documents")
    doc_headers = ["Teacher Name", "Document Type", "Document Name"]
    for col, h in enumerate(doc_headers, 1):
        style_header(ws2.cell(row=1, column=col, value=h))
    r = 2
    for t in teachers:
        for doc in all_docs.get(t['id'], []):
            alt = (r % 2 == 0)
            for col, v in enumerate([t.get('full_name'), doc.get('document_type'), doc.get('document_name')], 1):
                style_cell(ws2.cell(row=r, column=col, value=str(v or '')), alt=alt)
            r += 1
    auto_col_width(ws2)

    fname = f"Teachers_Report_{school_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return excel_response(wb, fname)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT 3 – FEE COLLECTION REPORT
# ═══════════════════════════════════════════════════════════════════════════════

@reports_bp.route('/reports/fees/export')
@login_required
@admin_required
def export_fees():
    school_id = request.args.get('school_id') or session.get('active_school_id') or session.get('school_id')
    class_id  = request.args.get('class_id')
    month     = request.args.get('month')     # e.g. "5"
    year      = request.args.get('year')      # e.g. "2025"

    conn = get_db()
    c    = conn.cursor()

    c.execute("SELECT name FROM schools WHERE id=%s", (school_id,))
    school_row  = c.fetchone()
    school_name = school_row[0] if school_row else "School"

    query = """
        SELECT fc.*,
               s.full_name AS student_name, s.student_code,
               cl.class_name, cl.section,
               u.full_name AS collector_name
        FROM fee_collections fc
        JOIN students s  ON fc.student_id = s.id
        LEFT JOIN classes cl  ON fc.class_id   = cl.id
        LEFT JOIN users u     ON fc.collected_by = u.id
        WHERE fc.school_id = %s
    """
    params = [school_id]

    if class_id:
        query += " AND fc.class_id = %s"
        params.append(class_id)
    if month:
        query += " AND fc.month = %s"
        params.append(month)
    if year:
        query += " AND fc.year = %s"
        params.append(year)

    query += " ORDER BY cl.class_name, s.full_name, fc.year, fc.month"

    print(f"DEBUG school_id={school_id} class_id={class_id} month={month} year={year}")
    print(f"DEBUG params={params}")
    c.execute(query, params)
    fees = fetchall_dict(c)
    print(f"DEBUG fees found={len(fees)}")
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fee Collections"

    period = ""
    if month and year:
        try:
            import calendar
            period = f" – {calendar.month_name[int(month)]} {year}"
        except Exception:
            period = f" – {month}/{year}"

    ws.merge_cells("A1:K1")
    ws["A1"] = f"{school_name} – Fee Collection Report{period}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Generated: {datetime.datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws["A2"].font = NORMAL_FONT
    ws["A2"].alignment = CENTER

    headers = ["#", "Receipt No.", "Student Code", "Student Name",
               "Class", "Section", "Month", "Year",
               "Amount (PKR)", "Payment Mode", "Collected By"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=4, column=col, value=h))

    total = 0
    for i, f in enumerate(fees, 1):
        row = i + 4
        alt = (i % 2 == 0)
        amount = f.get('amount', 0) or 0
        total += float(amount)
        vals = [
            i,
            f.get('receipt_number', ''),
            f.get('student_code', ''),
            f.get('student_name', ''),
            f.get('class_name', ''),
            f.get('section', ''),
            f.get('month', ''),
            f.get('year', ''),
            amount,
            f.get('payment_mode', ''),
            f.get('collector_name', ''),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            style_cell(cell, alt=alt)
            if col == 9:
                cell.number_format = '#,##0.00'

    # Total row
    total_row = len(fees) + 5
    ws.cell(row=total_row, column=8, value="TOTAL").font = BOLD_FONT
    total_cell = ws.cell(row=total_row, column=9, value=total)
    total_cell.font      = BOLD_FONT
    total_cell.fill      = PatternFill("solid", start_color="FFF2CC")
    total_cell.number_format = '#,##0.00'
    total_cell.border    = BORDER

    auto_col_width(ws)

    fname = f"Fee_Report_{school_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return excel_response(wb, fname)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT 4 – TEACHER SALARY REPORT
# ═══════════════════════════════════════════════════════════════════════════════

@reports_bp.route('/reports/salary/export')
@login_required
@admin_required
def export_salary():
    school_id  = request.args.get('school_id') or session.get('active_school_id') or session.get('school_id')
    teacher_id = request.args.get('teacher_id')   # optional: specific teacher
    month      = request.args.get('month')         # e.g. "5"
    year       = request.args.get('year')          # e.g. "2025"

    conn = get_db()
    c    = conn.cursor()

    c.execute("SELECT name FROM schools WHERE id=%s", (school_id,))
    school_row  = c.fetchone()
    school_name = school_row[0] if school_row else "School"

    # Check if teacher_salary_payments table exists; fall back to teachers.salary
    try:
        query = """
            SELECT tsp.*,
                   t.full_name AS teacher_name, t.teacher_code,
                   t.subject_specialization, t.salary AS base_salary,
                   u.full_name AS paid_by_name
            FROM teacher_salary_payments tsp
            JOIN teachers t ON tsp.teacher_id = t.id
            LEFT JOIN users u ON tsp.paid_by = u.id
            WHERE tsp.school_id = %s
        """
        params = [school_id]
        if teacher_id:
            query += " AND tsp.teacher_id = %s"
            params.append(teacher_id)
        if month:
            query += " AND tsp.month = %s"
            params.append(month)
        if year:
            query += " AND tsp.year = %s"
            params.append(year)
        query += " ORDER BY t.full_name, tsp.year, tsp.month"
        c.execute(query, params)
        salary_rows = fetchall_dict(c)
        mode = "payments"
    except Exception:
        # Fallback: use teachers.salary directly
        salary_rows = []
        mode = "base"

    if mode == "base" or not salary_rows:
        # Just show teacher base salaries
        q = "SELECT * FROM teachers WHERE school_id=%s"
        params = [school_id]
        if teacher_id:
            q += " AND id=%s"
            params.append(teacher_id)
        q += " ORDER BY full_name"
        c.execute(q, params)
        teachers = fetchall_dict(c)
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Teacher Salaries"

        period = ""
        if month and year:
            try:
                import calendar
                period = f" – {calendar.month_name[int(month)]} {year}"
            except Exception:
                period = f" – {month}/{year}"

        ws.merge_cells("A1:G1")
        ws["A1"] = f"{school_name} – Teacher Salary Report{period}"
        ws["A1"].font  = TITLE_FONT
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Generated: {datetime.datetime.now().strftime('%d-%b-%Y %H:%M')}"
        ws["A2"].font  = NORMAL_FONT
        ws["A2"].alignment = CENTER

        headers = ["#", "Teacher Code", "Full Name", "Specialization",
                   "Qualification", "Joining Date", "Monthly Salary (PKR)"]
        for col, h in enumerate(headers, 1):
            style_header(ws.cell(row=4, column=col, value=h))

        total = 0
        for i, t in enumerate(teachers, 1):
            row = i + 4
            alt = (i % 2 == 0)
            sal = t.get('salary', 0) or 0
            try:
                sal = float(sal)
            except Exception:
                sal = 0
            total += sal
            vals = [i, t.get('teacher_code',''), t.get('full_name',''),
                    t.get('subject_specialization',''), t.get('qualification',''),
                    str(t.get('joining_date','') or ''), sal]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                style_cell(cell, alt=alt)
                if col == 7:
                    cell.number_format = '#,##0.00'

        total_row = len(teachers) + 5
        ws.cell(row=total_row, column=6, value="TOTAL").font = BOLD_FONT
        tc = ws.cell(row=total_row, column=7, value=total)
        tc.font = BOLD_FONT
        tc.fill = PatternFill("solid", start_color="FFF2CC")
        tc.number_format = '#,##0.00'
        tc.border = BORDER

        auto_col_width(ws)

    else:
        conn.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary Payments"

        period = ""
        if month and year:
            try:
                import calendar
                period = f" – {calendar.month_name[int(month)]} {year}"
            except Exception:
                period = f" – {month}/{year}"

        ws.merge_cells("A1:J1")
        ws["A1"] = f"{school_name} – Teacher Salary Payment Report{period}"
        ws["A1"].font  = TITLE_FONT
        ws["A1"].alignment = CENTER

        ws.merge_cells("A2:J2")
        ws["A2"] = f"Generated: {datetime.datetime.now().strftime('%d-%b-%Y %H:%M')}"
        ws["A2"].font  = NORMAL_FONT
        ws["A2"].alignment = CENTER

        headers = ["#", "Teacher Code", "Teacher Name", "Specialization",
                   "Month", "Year", "Base Salary", "Paid Amount",
                   "Payment Mode", "Paid By"]
        for col, h in enumerate(headers, 1):
            style_header(ws.cell(row=4, column=col, value=h))

        total_paid = 0
        for i, r in enumerate(salary_rows, 1):
            row = i + 4
            alt = (i % 2 == 0)
            paid = float(r.get('paid_amount', 0) or 0)
            total_paid += paid
            vals = [i, r.get('teacher_code',''), r.get('teacher_name',''),
                    r.get('subject_specialization',''),
                    r.get('month',''), r.get('year',''),
                    r.get('base_salary',''), paid,
                    r.get('payment_mode',''), r.get('paid_by_name','')]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                style_cell(cell, alt=alt)
                if col in (7, 8):
                    cell.number_format = '#,##0.00'

        total_row = len(salary_rows) + 5
        ws.cell(row=total_row, column=7, value="TOTAL").font = BOLD_FONT
        tc = ws.cell(row=total_row, column=8, value=total_paid)
        tc.font = BOLD_FONT
        tc.fill = PatternFill("solid", start_color="FFF2CC")
        tc.number_format = '#,##0.00'
        tc.border = BORDER

        auto_col_width(ws)

    fname = f"Salary_Report_{school_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return excel_response(wb, fname)
